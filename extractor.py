"""기성고 검사요청서/보고서 PDF·Excel에서 표준 항목을 추출한다.

문서는 고정 양식(재료비/노무비/경비/일반관리비/이윤 ...)이므로,
행은 키워드로, 열은 헤더 위치로 매칭한다. 텍스트 레이어가 있으면
pdfplumber로 그대로 추출하고, 스캔 이미지라 텍스트가 없으면
OCR(Tesseract)로 대체한다.
"""
import io
import os
import re
import shutil

import pdfplumber
import pymupdf
import pytesseract
from openpyxl import load_workbook

TESSERACT_CMD = (
    os.environ.get("TESSERACT_CMD")
    or shutil.which("tesseract")
    or r"C:\Program Files\Tesseract-OCR\tesseract.exe"
)
pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

COLUMNS = ["계약금액", "전회누계", "기성율", "금회누계금액", "금회지급액", "비고"]

# (key, 표시라벨, 매칭키워드, 핵심필드여부[이슈#2: 건강보험료/연금보험료/퇴직공제부금])
ROWS = [
    ("material_direct", "1) 직접재료비", "직접재료비", False),
    ("material_indirect", "2) 간접재료비", "간접재료비", False),
    ("material_subtotal", "소계 (재료비)", "소계", False),
    ("labor_direct", "1) 직접노무비", "직접노무비", False),
    ("labor_indirect", "2) 간접노무비", "간접노무비", False),
    ("labor_subtotal", "소계 (노무비)", "소계", False),
    ("exp_industrial_accident", "1) 산재보험료", "산재보험료", False),
    ("exp_employment", "2) 고용보험료", "고용보험료", False),
    ("exp_health", "3) 건강보험료", "건강보험료", True),
    ("exp_pension", "4) 연금보험료", "연금보험료", True),
    ("exp_longterm_care", "5) 노인장기요양보험료", "노인장기요양보험료", False),
    ("exp_retirement_mutual_aid", "6) 퇴직공제부금", "퇴직공제부금", True),
    ("exp_safety_health_mgmt", "7) 산업안전보건관리비", "산업안전보건관리비", False),
    ("exp_safety_mgmt", "8) 안전관리비", "안전관리비", False),
    ("exp_environment", "9) 환경보전비", "환경보전비", False),
    ("exp_other", "10) 기타경비", "기타경비", False),
    ("exp_performance_bond", "11) 공사이행보증수수료", "공사이행보증수수료", False),
    ("exp_subcontract_bond", "12) 건설하도급대금지급보증서 발급수수료", "건설하도급대금지급", False),
    ("exp_equipment_bond", "13) 건설기계대여금지급보증서 발급수수료", "건설기계대여", False),
    ("exp_subtotal", "소계 (경비)", "소계", False),
    ("general_mgmt", "4. 일반관리비", "일반관리비", False),
    ("profit", "5. 이윤", "이윤", False),
    ("total", "총계", "총계", False),
    ("rounding", "단수정리", "단수정리", False),
]

GROUP_MARKERS = [
    ("재료비", "material_subtotal"),
    ("노무비", "labor_subtotal"),
    ("경비", "exp_subtotal"),
]

META_PATTERNS = {
    "calc_period": r"기성고\s*산정기간\D*([\d.\t ]+~[\d.\t ]+)",
    "project_name": r"공\s*사\s*명\s*[:：]\s*([^\n]+?)(?:계\s*약\s*번\s*호|$)",
    "contract_no": r"계\s*약\s*번\s*호\s*[:：]\s*(\S+)",
    "project_period": r"공사기간\s*[:：]\s*([\d.\t ]+~[\d.\t ]+)",
    "supply_amount": r"공\s*급\s*가\s*액\s*[:：]\s*([\d,\t ]+)",
    "contract_date": r"계\s*약\s*일\s*[:：]\s*([\d.\t ]+)",
    "vat": r"부\s*가\s*세\s*[:：]\s*([\d,\t ]+)",
    "contractor": r"계\s*약\s*자\s*[:：]\s*([^\n]+?)(?:합\s*계|$)",
}


def norm(s):
    return re.sub(r"\s+", "", s or "")


def clean_amount(s):
    if s is None:
        return None
    s = str(s).replace("\n", " ").strip()
    s = s.replace("₩", "").replace(",", "").replace(" ", "")
    if s in ("", "-", "None"):
        return 0
    try:
        return int(s)
    except ValueError:
        m = re.search(r"-?\d+", s)
        return int(m.group()) if m else None


def clean_percent(s):
    if s is None:
        return None
    s = str(s).replace("%", "").strip()
    if s == "":
        return None
    try:
        return round(float(s), 3)
    except ValueError:
        return None


def _match_row_key(cell_text, current_group_idx):
    label = norm(cell_text)
    if not label:
        return None, current_group_idx
    for i, (group_name, _) in enumerate(GROUP_MARKERS):
        if label.startswith(f"{i + 1}.{group_name}") or label == group_name:
            return None, i
    for key, _, keyword, _ in ROWS:
        if key.endswith("subtotal"):
            continue
        if keyword in label:
            return key, current_group_idx
    if "소계" in label:
        if 0 <= current_group_idx < len(GROUP_MARKERS):
            return GROUP_MARKERS[current_group_idx][1], current_group_idx
    return None, current_group_idx


def _parse_meta_text(text):
    meta = {}
    for key, pattern in META_PATTERNS.items():
        m = re.search(pattern, text, re.S)
        if m:
            meta[key] = re.sub(r"\s+", " ", m.group(1)).strip()
    return meta


# pdfplumber가 이 양식에서 뽑아내는 표는 열 위치가 항상 고정이다:
# 0:구분 1:계약금액 2:전회누계[A] 3:기성율 4:금회누계금액[B] 5:금회지급액[C] 6:(병합, 빈칸) 7:비고
PDF_TABLE_COLUMN_COUNT = 8


def _parse_row_by_position(cells):
    def at(i):
        return cells[i] if i < len(cells) else None

    return {
        "계약금액": clean_amount(at(1)),
        "전회누계": clean_amount(at(2)),
        "기성율": clean_percent(at(3)),
        "금회누계금액": clean_amount(at(4)),
        "금회지급액": clean_amount(at(5)),
        "비고": str(at(7) or "").replace("\n", " ").strip(),
    }


def _parse_row_heuristic(cells):
    numeric_cells = [c for c in cells[1:] if str(c).strip() not in ("", "None")]
    contract_amount = clean_amount(numeric_cells[0]) if len(numeric_cells) > 0 else None
    prev_cum = clean_amount(numeric_cells[1]) if len(numeric_cells) > 1 else None

    rate = None
    curr_cum = None
    payment = None
    note = ""
    for val in numeric_cells[2:]:
        sval = str(val).strip()
        if "%" in sval and rate is None:
            rate = clean_percent(sval)
        elif re.search(r"\d", sval) and re.match(r"^[₩\-\d,\.\s%]+$", sval):
            if curr_cum is None:
                curr_cum = clean_amount(sval)
            elif payment is None:
                payment = clean_amount(sval)
        else:
            note = (note + " " + sval).strip()
    return {
        "계약금액": contract_amount,
        "전회누계": prev_cum,
        "기성율": rate,
        "금회누계금액": curr_cum,
        "금회지급액": payment,
        "비고": note,
    }


def _extract_table_rows(table_rows, fixed_columns=False):
    """pdfplumber/openpyxl 스타일 2차원 리스트에서 표준 항목 dict를 만든다.

    fixed_columns=True면 이 양식 고유의 고정 열 위치로 파싱한다(더 정확함).
    그렇지 않으면(예: 열 배치가 다를 수 있는 엑셀) 값의 순서로 추정한다.
    """
    items = {}
    group_idx = -1
    for row in table_rows:
        if not row:
            continue
        cells = [c if c is not None else "" for c in row]
        first = str(cells[0])
        key, group_idx = _match_row_key(first, group_idx)
        if key is None:
            continue
        if fixed_columns and len(cells) >= 6:
            items[key] = _parse_row_by_position(cells)
        else:
            items[key] = _parse_row_heuristic(cells)
    return items


def _rows_from_pdfplumber_table(table):
    return _extract_table_rows(table, fixed_columns=True)


def _ocr_page_to_text(page):
    pix = page.get_pixmap(dpi=300)
    img_bytes = pix.tobytes("png")
    from PIL import Image

    img = Image.open(io.BytesIO(img_bytes))
    return pytesseract.image_to_string(img, lang="kor+eng")


def _extract_items_from_ocr_text(text):
    """스캔 이미지라 표 구조가 깨진 경우, 키워드 주변의 숫자열을 정규식으로 회수한다."""
    items = {}
    lines = text.splitlines()
    group_idx = -1
    for line in lines:
        key, group_idx = _match_row_key(line, group_idx)
        if key is None:
            continue
        nums = re.findall(r"[₩\\]?\s?[\d][\d,]{2,}", line)
        cleaned = [clean_amount(n) for n in nums]
        cleaned = [n for n in cleaned if n is not None]
        pct_m = re.search(r"(\d{1,3}(?:\.\d+)?)\s?%", line)
        items[key] = {
            "계약금액": cleaned[0] if len(cleaned) > 0 else None,
            "전회누계": cleaned[1] if len(cleaned) > 1 else None,
            "기성율": clean_percent(pct_m.group(1)) if pct_m else None,
            "금회누계금액": cleaned[2] if len(cleaned) > 2 else None,
            "금회지급액": cleaned[3] if len(cleaned) > 3 else None,
            "비고": "",
        }
    return items


def _is_summary_page_text(text):
    label = norm(text)
    return "계약금액" in label and ("구분" in label or "기성고" in label)


def extract_from_pdf(path):
    """기성고 요약표가 있는 단일 페이지를 찾아 추출한다.

    문서 번들에는 요약표와 무관한 첨부(원장/증명서 등) 페이지가 다수
    섞여 있을 수 있어, 페이지별로 후보를 모은 뒤 가장 많은 항목을
    회수한 페이지 하나만 채택한다 (여러 페이지를 그대로 병합하면
    다른 표의 값이 요약표 값을 덮어써 버린다).
    """
    native_candidates = []
    ocr_pages = []

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if len(text.strip()) > 50:
                if _is_summary_page_text(text):
                    for table in page.extract_tables():
                        parsed = _rows_from_pdfplumber_table(table)
                        if parsed:
                            native_candidates.append((len(parsed), parsed, text))
            else:
                ocr_pages.append(page.page_number - 1)

    if native_candidates:
        native_candidates.sort(key=lambda c: c[0], reverse=True)
        count, items, text = native_candidates[0]
        return {"meta": _parse_meta_text(text), "items": items, "source": "native"}

    if ocr_pages:
        doc = pymupdf.open(path)
        ocr_candidates = []
        for page_index in ocr_pages[:10]:
            text = _ocr_page_to_text(doc[page_index])
            parsed = _extract_items_from_ocr_text(text)
            if parsed:
                ocr_candidates.append((len(parsed), parsed, text))
        doc.close()
        if ocr_candidates:
            ocr_candidates.sort(key=lambda c: c[0], reverse=True)
            count, items, text = ocr_candidates[0]
            return {"meta": _parse_meta_text(text), "items": items, "source": "ocr"}

    return {"meta": {}, "items": {}, "source": "none"}


def _find_header_row(rows, max_scan=10):
    for idx, row in enumerate(rows[:max_scan]):
        joined = norm(" ".join(str(c) for c in row if c is not None))
        if "계약금액" in joined:
            return idx
    return None


def extract_from_excel(path):
    wb = load_workbook(path, data_only=True)
    all_items = {}
    meta = {}
    for ws in wb.worksheets:
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        full_text = "\n".join(
            " ".join(str(c) for c in row if c is not None) for row in rows
        )
        meta.update(_parse_meta_text(full_text))
        header_idx = _find_header_row(rows)
        data_rows = rows[header_idx + 1:] if header_idx is not None else rows
        parsed = _extract_table_rows(data_rows)
        all_items.update(parsed)
    return {"meta": meta, "items": all_items, "source": "excel"}


def extract(path, filename):
    ext = filename.lower().rsplit(".", 1)[-1]
    if ext == "pdf":
        return extract_from_pdf(path)
    if ext in ("xlsx", "xlsm", "xls"):
        return extract_from_excel(path)
    raise ValueError(f"지원하지 않는 파일 형식입니다: {ext}")


def row_definitions():
    return ROWS
