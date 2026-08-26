"""하도급사별로 양식이 다른 "노무비 지급 명세서" 엑셀을 공통 스키마로 정규화한다 (이슈 #5-3).

담당자 의견 3번은 "하도급사 제출 양식이 상이하므로 하나로 통일 필요 - 자체 template
구성필요"였다. 그런데 실제 제출본 3사(신보/신한/NSC, 26.03·26.04)를 열어보니
**3사 모두 이미 일자별 출역(공수)을 갖고 있고 배치만 다르다**. 그래서 새 양식을
강요하기 전에 3종을 자동 인식해 공통 스키마로 정규화하는 쪽이 빠르고, 그 공통
스키마를 엑셀로 내보낸 것이 곧 배포할 "표준 템플릿"이 된다.

## 실측한 3사 양식의 공통 구조와 차이

세 양식 모두 **한 사람이 2행**이고, 위 행이 전반기(1~15일), 아래 행이 후반기
(16~말일) 공수다. 라벨도 2단이라 위 행에는 1단 라벨(출력일수/일급/지급총액...),
아래 행에는 2단 라벨(출력공수/지방소득세/실지급액...)의 값이 들어간다.

| | 신보 · 신한 | NSC |
|---|---|---|
| 시트 | `노무비명세서(...)` / `지급명세서` | `지급명세서` |
| 날짜 칸 | `datetime` 값 (예: 2026-03-01) | 일 번호 문자열 (`'1'`~`'15'`) |
| 라벨행 / 날짜행 | 같은 행이 겸함 (좌측=라벨, 우측=날짜) | 라벨행과 날짜행이 **교차** (r3 라벨, r4 날짜, r5 라벨, r6 날짜) |
| 헤더 반복 | 1회 | **5회** (페이지마다 반복 + 합계 행) |
| 업체명 | 시트 안에 있음 (`작 성 자 :` / `상     호 :`) | 시트에 없음 → 파일명에서 |
| 연월 | 날짜 칸이 datetime이라 그대로 | 제목 `2026년 4월 ...` 에서 |

그래서 행 위치를 고정하지 않고 (1) 성명 헤더가 나오는 행마다 "헤더 구간"을 열고,
(2) 그 구간에서 날짜 칸이 10개 이상인 행을 찾아 컬럼->날짜 표를 만들고, (3) 성명
칸에 사람 이름이 있는 행을 블록 시작으로 잡는 방식으로 처리한다 -- NSC처럼 헤더가
중간에 다시 나와도, 신보처럼 뒤에 빈 행이 수백 개 붙어 있어도 같은 코드로 읽힌다.

## 검증 지표

세 양식 모두 인별 "출력공수"(NSC는 "공수") 칸에 총공수가 적혀 있다. 일자별 칸을
합한 값과 이 기재값이 다르면 `needs_review`로 표시한다 -- 병합셀/열 밀림을 조용히
넘기지 않기 위한 내부 일관성 검증이다(건강보험 "고지x2=납부"와 같은 패턴).
"""
import datetime
import re

from openpyxl import Workbook, load_workbook

# 합계 행이 성명 칸에 걸리는 경우가 있어(NSC 실측: 163/189행) 사람 이름과 구분한다.
TOTAL_LABELS = {"계", "합계", "소계", "총계", "누계", "이월"}
NAME_RE = re.compile(r"^[가-힣]{2,5}$")
BIRTH6_RE = re.compile(r"^(\d{6})\s*-")
TITLE_YM_RE = re.compile(r"(\d{4})\s*년\s*(\d{1,2})\s*월")
FILE_YM_RE = re.compile(r"(\d{2})[.\-](\d{2})\s*월")
MIN_DAY_CELLS = 10  # 이만큼 날짜 칸이 있으면 "날짜행"으로 본다

# 양식마다 라벨 표기가 달라(일급/임금단가, 차감지급액/실지급액) 정규 이름으로 모은다.
CANONICAL_LABELS = {
    "출력일수": "출역일수_기재",
    "출력공수": "총공수_기재",
    "공수": "총공수_기재",
    "능률공수": "능률공수",
    "일급": "일급",
    "임금단가": "일급",
    "지급총액": "지급총액",
    "노무비총액": "지급총액",
    "차감지급액": "실지급액",
    "실지급액": "실지급액",
}

COMPANY_LABELS = ("작성자", "상호", "업체명", "회사명", "제출자")
SITE_LABELS = ("공사명", "현장명", "사업장명")


def norm(value):
    return re.sub(r"\s+", "", str(value or ""))


def _as_day(value):
    """날짜 칸 값을 1~31 사이 일(day) 번호로 해석한다(아니면 None)."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and float(value).is_integer() and 1 <= value <= 31:
        return int(value)
    if isinstance(value, str):
        text = value.strip()
        if text.isdigit() and 1 <= int(text) <= 31:
            return int(text)
    return None


def _is_day_row(row):
    """날짜(1~15일 / 16~말일) 헤더 행인지 판별한다.

    개수만 세면 안 된다 -- 사람 행의 공수가 대부분 `1`이라 "일 번호 1"로 보여
    사람 행까지 날짜 행으로 잡힌다(실측: 신보 시트에서 r6~r13이 전부 오인됨).
    진짜 날짜 행은 컬럼 순서대로 값이 **단조 증가**한다는 점으로 가른다.
    """
    sequence = [
        v if isinstance(v, datetime.datetime) else _as_day(v)
        for v in row
        if isinstance(v, datetime.datetime) or _as_day(v) is not None
    ]
    if len(sequence) < MIN_DAY_CELLS:
        return False
    return all(b > a for a, b in zip(sequence, sequence[1:]))


def _as_manday(value):
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value else None
    text = str(value).strip()
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number or None


def _labelled_value(rows, labels, max_scan=6):
    """"공 사 명 : (값)" 또는 "공 사 명 :" | "(값)" 두 형태를 모두 읽는다."""
    for row in rows[:max_scan]:
        for idx, cell in enumerate(row):
            if not isinstance(cell, str):
                continue
            key = norm(cell).rstrip(":：")
            if key not in labels:
                continue
            after = re.sub(r"^[^:：]*[:：]\s*", "", cell).strip()
            if after and norm(after) != key:
                return after
            for following in row[idx + 1:]:
                # 값 칸이 비어 있고 옆에 또 다른 라벨만 있는 경우(NSC의 "현 장 명")
                # 라벨을 값으로 착각하지 않도록 라벨과 같은 글자면 버린다.
                if isinstance(following, str) and following.strip() and norm(following) != key:
                    return following.strip()
    return None


def _year_month(rows, day_dates, filename):
    for date in day_dates:
        if isinstance(date, datetime.datetime):
            return f"{date.year:04d}-{date.month:02d}"
    for row in rows[:6]:
        for cell in row:
            if isinstance(cell, str):
                m = TITLE_YM_RE.search(cell)
                if m:
                    return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    m = FILE_YM_RE.search(filename or "")
    if m:
        return f"20{m.group(1)}-{m.group(2)}"
    return None


def _find_name_column(rows):
    for row in rows[:12]:
        for idx, cell in enumerate(row):
            if norm(cell) == "성명":
                return idx
    return None


def _header_rows(rows):
    return [i for i, row in enumerate(rows) if any(norm(v) == "성명" for v in row)]


def _section_layout(rows, header_idx, limit):
    """헤더 구간에서 (컬럼->일/날짜) 표 2개와 1단/2단 라벨 컬럼을 뽑는다."""
    day_rows = [i for i in range(header_idx, min(header_idx + 8, limit)) if _is_day_row(rows[i])]
    if not day_rows:
        return None

    def day_map(row):
        out = {}
        for col, value in enumerate(row):
            if isinstance(value, datetime.datetime):
                out[col] = value
            else:
                day = _as_day(value)
                if day is not None:
                    out[col] = day
        return out

    first_half = day_map(rows[day_rows[0]])
    second_half = day_map(rows[day_rows[1]]) if len(day_rows) > 1 else {}

    def labels_in(row):
        out = {}
        for col, value in enumerate(row):
            if not isinstance(value, str) or _as_day(value) is not None:
                continue
            key = CANONICAL_LABELS.get(norm(value))
            if key:
                out[key] = col
        return out

    tier_a = labels_in(rows[header_idx])
    tier_b = {}
    for i in range(header_idx + 1, day_rows[-1] + 1):
        for key, col in labels_in(rows[i]).items():
            tier_b.setdefault(key, col)

    return {
        "first_half": first_half,
        "second_half": second_half,
        "tier_a": tier_a,
        "tier_b": tier_b,
        "body_start": day_rows[-1] + 1,
        "body_end": limit,
    }


def _resolve_date(value, year_month):
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if not year_month:
        return None
    return f"{year_month}-{int(value):02d}"


def _person_identity(row_a, row_b, name_col, first_date_col):
    name = row_a[name_col] if name_col < len(row_a) else None
    name = name.strip() if isinstance(name, str) else None
    if not name or not NAME_RE.match(name) or norm(name) in TOTAL_LABELS:
        return None, None, None

    birth6 = None
    job = None
    for row in (row_a, row_b):
        for col in range(min(first_date_col, len(row))):
            if col == name_col:
                continue
            cell = row[col]
            if not isinstance(cell, str):
                continue
            text = cell.strip()
            m = BIRTH6_RE.match(text)
            if m and birth6 is None:
                birth6 = m.group(1)
            elif job is None and re.search(r"[가-힣]", text) and not m:
                # 직종/팀명 (신보 "전공", 신한 "판넬 1팀", NSC "배관공")
                job = text
    return name, birth6, job


def parse_labor_ledger(path, filename=None):
    """노무비 지급 명세서 엑셀 하나를 읽어 (일자별 출역 레코드, 인별 요약)을 돌려준다."""
    filename = filename or str(path)
    workbook = load_workbook(path, data_only=True)
    attendance = []
    summaries = []

    try:
        for sheet_name in workbook.sheetnames:
            rows = [list(r) for r in workbook[sheet_name].iter_rows(values_only=True)]
            name_col = _find_name_column(rows)
            headers = _header_rows(rows)
            if name_col is None or not headers:
                continue  # 명세서 표가 없는 시트(내역서/요약 등)

            company = _labelled_value(rows, COMPANY_LABELS) or _company_from_filename(filename)
            site = _labelled_value(rows, SITE_LABELS)

            for order, header_idx in enumerate(headers):
                limit = headers[order + 1] if order + 1 < len(headers) else len(rows)
                layout = _section_layout(rows, header_idx, limit)
                if not layout:
                    continue

                day_values = [*layout["first_half"].values(), *layout["second_half"].values()]
                year_month = _year_month(rows, day_values, filename)
                first_date_col = min([*layout["first_half"], *layout["second_half"]], default=name_col + 1)

                i = layout["body_start"]
                while i + 1 < layout["body_end"]:
                    row_a, row_b = rows[i], rows[i + 1]
                    name, birth6, job = _person_identity(row_a, row_b, name_col, first_date_col)
                    if not name:
                        i += 1
                        continue

                    base = {
                        "company": company,
                        "site": site,
                        "person": name,
                        "birth6": birth6,
                        "job": job,
                        "year_month": year_month,
                    }
                    days = _collect_days(row_a, row_b, layout, year_month)
                    for date, manday in days:
                        attendance.append({**base, "date": date, "manday": manday})

                    summaries.append(
                        {
                            **base,
                            "출역일수_계산": len(days),
                            "총공수_계산": round(sum(m for _, m in days), 2),
                            **_summary_values(row_a, row_b, layout),
                            "source_sheet": sheet_name,
                            "source_file": filename,
                        }
                    )
                    i += 2
    finally:
        workbook.close()

    for summary in summaries:
        stated = summary.get("총공수_기재")
        summary["needs_review"] = (
            stated is not None and abs(float(stated) - summary["총공수_계산"]) > 0.01
        )
        # 일급 x 총공수를 그대로 곱해 둔다 -- 신보/신한은 이 값이 지급총액과 정확히
        # 일치하지만(실측 95/95, 86/86, 63/63, 68/68), NSC는 여기에 "능률공수"가
        # 더해져 일치하지 않는다(실측 67/71 불일치, 예: 25공수+5능률=30x일급).
        # 양식마다 가산 규칙이 달라 자동 판정하지 않고 검산값만 같이 보여준다.
        rate = summary.get("일급")
        if rate:
            summary["지급총액_검산"] = round(rate * summary["총공수_계산"], 0)
    return attendance, summaries


def _collect_days(row_a, row_b, layout, year_month):
    days = []
    for row, day_map in ((row_a, layout["first_half"]), (row_b, layout["second_half"])):
        for col, day_value in day_map.items():
            manday = _as_manday(row[col]) if col < len(row) else None
            if manday is None:
                continue
            date = _resolve_date(day_value, year_month)
            if date:
                days.append((date, manday))
    days.sort()
    return days


def _summary_values(row_a, row_b, layout):
    out = {}
    for tier, row in (("tier_a", row_a), ("tier_b", row_b)):
        for key, col in layout[tier].items():
            if col < len(row) and isinstance(row[col], (int, float)) and not isinstance(row[col], bool):
                out.setdefault(key, row[col])
    return out


def _company_from_filename(filename):
    """NSC처럼 시트 안에 업체명이 없는 양식은 파일명 앞부분에서 가져온다
    (예: "NSC_26.04월 노무비 지급명세서.xlsx" -> "NSC")."""
    stem = re.sub(r"\.[A-Za-z]+$", "", str(filename or "").split("/")[-1].split("\\")[-1])
    head = re.split(r"[_(]", stem, maxsplit=1)[0].strip()
    return head or None


# ---------------------------------------------------------------------------
# 표준 템플릿 (통일 양식) 내보내기
# ---------------------------------------------------------------------------

ATTENDANCE_HEADERS = ["업체명", "현장명", "성명", "생년월일6", "직종", "연월", "일자", "공수"]
SUMMARY_HEADERS = [
    ("업체명", "company"),
    ("현장명", "site"),
    ("성명", "person"),
    ("생년월일6", "birth6"),
    ("직종", "job"),
    ("연월", "year_month"),
    ("출역일수(계산)", "출역일수_계산"),
    ("총공수(계산)", "총공수_계산"),
    ("출역일수(기재)", "출역일수_기재"),
    ("총공수(기재)", "총공수_기재"),
    ("능률공수", "능률공수"),
    ("일급", "일급"),
    ("지급총액", "지급총액"),
    ("지급총액 검산(일급x총공수)", "지급총액_검산"),
    ("실지급액", "실지급액"),
    ("확인필요", "needs_review"),
    ("원본파일", "source_file"),
]


def build_labor_template_workbook(attendance, summaries):
    """정규화 결과를 그대로 표준 양식으로 내보낸다.

    "출역" 시트는 (사람 x 일자) 한 행씩이라 다른 자료(출입기록/TBM/인부명부)와
    바로 대조할 수 있는 형태이고(이슈 #5-4의 입력), "월별 출역표" 시트는 하도급사가
    지금 쓰는 것과 같은 (사람 x 1~31일) 격자라 그대로 배포용 템플릿이 된다.
    """
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "출역"
    sheet.append(ATTENDANCE_HEADERS)
    for row in sorted(attendance, key=lambda r: (r["company"] or "", r["person"] or "", r["date"])):
        sheet.append(
            [row["company"], row["site"], row["person"], row["birth6"], row["job"],
             row["year_month"], row["date"], row["manday"]]
        )

    summary_sheet = workbook.create_sheet("인별 요약")
    summary_sheet.append([label for label, _ in SUMMARY_HEADERS])
    for row in sorted(summaries, key=lambda r: (r["company"] or "", r["person"] or "")):
        summary_sheet.append([_summary_cell(row, key) for _, key in SUMMARY_HEADERS])

    _append_month_grid(workbook, attendance)
    return workbook


def _summary_cell(row, key):
    if key == "needs_review":
        return "확인필요" if row.get(key) else ""
    return row.get(key)


def _append_month_grid(workbook, attendance):
    """하도급사 배포용 (사람 x 1~31일) 격자. 지금 3사가 쓰는 모양과 같되 한 사람이
    2행이 아니라 1행이라 기계가 읽기 쉽다."""
    if not attendance:
        return
    sheet = workbook.create_sheet("월별 출역표")
    sheet.append(["업체명", "성명", "생년월일6", "직종", "연월", *[str(d) for d in range(1, 32)], "총공수"])

    grouped = {}
    for row in attendance:
        key = (row["company"], row["person"], row["birth6"], row["job"], row["year_month"])
        grouped.setdefault(key, {})[int(row["date"][-2:])] = row["manday"]

    for (company, person, birth6, job, year_month), by_day in sorted(
        grouped.items(), key=lambda kv: (kv[0][0] or "", kv[0][1] or "", kv[0][4] or "")
    ):
        sheet.append(
            [company, person, birth6, job, year_month,
             *[by_day.get(d) for d in range(1, 32)], round(sum(by_day.values()), 2)]
        )
