"""퇴직공제부금 OCR 추출 파이프라인 (이슈 #5 연장, 2026-08-24).

`sample/06-05. 제20회 기성 실적정산(퇴직공제).pdf`는 텍스트 레이어가 없는 순수
스캔본(전 페이지 chars=0, 이미지만 존재)이라 `extract_compare.py`의 나머지
문서 종류(health/pension)처럼 pdfplumber로 글자를 바로 못 읽는다. 그래서 이
파일만 별도 모듈로 뺐다:

1. pymupdf(fitz)로 각 페이지를 300dpi 이미지로 렌더링
2. 표 격자선(가로/세로 검은 선) 위치를 픽셀 분석으로 직접 찾음(OpenCV 없이
   numpy만으로 -- 이 프로젝트 venv에 opencv가 없어서 다크픽셀 합 방식을 씀)
3. 셀 하나씩 pytesseract를 부르면 314셀/페이지 x 188페이지가 되어 셀당 프로세스
   기동 비용(~150ms)만으로 페이지당 52초가 나온다(실측). 대신 "같은 컬럼의 모든
   행을 세로로 이어붙인 이미지 1장"을 컬럼당 1번만 OCR하고, 각 글자의 픽셀
   좌표를 우리가 이미 알고 있는 행 경계에 직접 매핑한다 -- 이러면 컬럼 9개
   기준 페이지당 OCR 호출이 9번으로 줄어 페이지당 ~6~7초가 된다(실측, 약 8배).
   줄바꿈 개수로 행을 나누지 않는 이유: 숫자 한 글자가 통째로 안 읽히면 그
   줄이 조용히 통째로 사라져 이후 모든 행이 한 칸씩 밀린다(실측 확인) --
   좌표 매핑은 실패한 칸만 빈 값이 되고 다른 행에 영향이 없다.

이 파일에서 다루는 표는 "OO년 OO월 공제부금 납부 신고 내역"(9열: NO/성명/
생년월일/근로년월/근로일수/직종/업체명/단위공사명/납부액) 표 하나뿐이다. 같은
PDF 안에 최소 두 종류가 더 섞여 있는데:
  - "공제부금납부확인서"(건설근로자공제회 발급, 원청 전체 합계 확인서, QR/직인)
  - "NO/성명/생년월일/근로년월/근로일수/전화번호/직종/국적/여권번호/입력자/
    상태구분" 11열 표(근로내역신고 원본, 금액 컬럼이 없어 대조에 못 씀)
둘 다 목적에 안 맞아 grid 검출 단계에서 자동으로 걸러내고 건너뛴다(아래
`detect_table_grid` 참고 -- 비율 템플릿에 안 맞으면 None).
"""
import re
import statistics
from pathlib import Path

import fitz  # pymupdf
import numpy as np
import openpyxl
import pandas as pd
import pytesseract
from PIL import Image
from pytesseract import Output

DPI = 300

# 9열(NO/성명/생년월일/근로년월/근로일수/직종/업체명/단위공사명/납부액) 경계선의
# 상대 위치(표 왼쪽 끝=0.0, 오른쪽 끝=1.0). 26.02월 표(35명, 페이지 온전한
# 격자선)에서 실측한 값 -- 페이지마다 스캔 위치가 몇 픽셀씩 흔들려도 표 전체
# 폭 대비 비율은 문서 생성 템플릿이 고정이라 안정적이다.
TABLE_RATIOS = [0.0, 0.0466, 0.1394, 0.2469, 0.3396, 0.4272, 0.5466, 0.7017, 0.8568, 1.0]

# 순번(NO)열은 OCR에 안 맡긴다 -- 이 표는 매 페이지 끊김 없이 1씩 증가하는
# 절대 순번이라 페이지 내 행 위치만으로 이미 알 수 있고, OCR로 읽으면 오히려
# 오탐(실측: "1"->"4" 오독) 위험만 생긴다.
COLUMN_SPECS = [
    ("성명", "kor", None),
    ("생년월일", "eng", "0123456789."),
    ("근로년월", "eng", "0123456789-"),
    ("근로일수", "eng", "-0123456789"),
    ("직종", "kor", None),
    ("업체명", "kor", None),
    ("단위공사명", "eng", "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-"),
    ("납부액", "eng", "-0123456789,"),
]

REPORT_MONTH_RE = re.compile(r"(20\d\d)\D{0,3}(\d{1,2})\D{0,3}월")
NOISE_LEADING_RE = re.compile(r"^[^가-힣A-Za-z0-9(]+")


def _render_page(page, dpi=DPI):
    pix = page.get_pixmap(dpi=dpi)
    return Image.frombytes("RGB", (pix.width, pix.height), pix.samples)


def _detect_lines(dark, axis, rng_a, rng_b, min_len_ratio):
    if axis == "v":
        region = dark[rng_a:rng_b, :]
        s = region.sum(axis=0)
        length = rng_b - rng_a
    else:
        region = dark[:, rng_a:rng_b]
        s = region.sum(axis=1)
        length = rng_b - rng_a
    thresh = min_len_ratio * length
    idx = np.where(s > thresh)[0]
    groups = []
    for i in idx:
        if groups and i - groups[-1][-1] <= 4:
            groups[-1].append(i)
        else:
            groups.append([i])
    return [int(np.mean(g)) for g in groups]


def detect_table_grid(gray_arr):
    """"납부 신고 내역" 9열 표만 골라 세로/가로 격자선 좌표를 돌려준다.

    이 PDF에는 형태가 다른 표(11열 근로내역신고, 공제부금납부확인서 등)와
    표가 아예 없는 페이지(표지, 기성현황 요약, 결재란)가 섞여 있다 --
    TABLE_RATIOS 템플릿과 위치가 맞는 세로선이 7개 미만이면(다른 표거나 표가
    없는 페이지) None을 돌려줘서 호출부가 그 페이지를 건너뛰게 한다.
    """
    dark = gray_arr < 190
    v_cand = _detect_lines(dark, "v", 250, 3300, 0.30)
    if len(v_cand) < 2:
        return None
    x0, x1 = v_cand[0], v_cand[-1]
    width = x1 - x0
    if width < 1500:
        return None
    expected = [x0 + r * width for r in TABLE_RATIOS]
    v_lines = []
    matched = 0
    for ex in expected:
        best = min(v_cand, key=lambda c: abs(c - ex))
        if abs(best - ex) <= 25:
            v_lines.append(best)
            matched += 1
        else:
            v_lines.append(int(ex))
    if matched < 7:
        return None
    h_lines = _detect_lines(dark, "h", x0, x1, 0.5)
    if len(h_lines) < 3:
        return None
    return v_lines, h_lines


def _darken_red_ink(pil_img, threshold=25):
    """정정(correction) 행은 빨간 글자로 인쇄된다. 실측(300dpi RGB 크롭):
    빨간 글자의 그레이스케일 변환값이 최대 208까지 나와(검은 글자보다 훨씬
    밝음) 흰 배경과 명암비가 약하고, 그 결과 pytesseract 내부 이진화가
    마이너스 부호나 자릿수를 통째로 놓치는 사례가 실측으로 확인됐다(예:
    근로일수 "-1"이 "1"로, 금액 "-6,500"이 "7 6,500 ,"로 읽힘). R채널이
    G/B보다 뚜렷이 높은 픽셀(빨간 글자)만 골라 검게 눌러주면 같은 셀에서
    부호가 정확히 복원된다(실측 확인) -- 검은 글자/흰 배경 픽셀은 그대로
    둔다."""
    arr = np.array(pil_img.convert("RGB")).astype(np.int16)
    r, g, b = arr[:, :, 0], arr[:, :, 1], arr[:, :, 2]
    mask = (r - g > threshold) & (r - b > threshold)
    arr[mask] = [0, 0, 0]
    return Image.fromarray(arr.astype(np.uint8))


def _detect_report_month(img, top_y):
    crop = img.crop((0, 0, img.width, max(top_y, 1)))
    text = pytesseract.image_to_string(crop, lang="kor", config="--psm 6")
    m = REPORT_MONTH_RE.search(text.replace("\n", " "))
    if not m:
        return None
    year, month = int(m.group(1)), int(m.group(2))
    return f"{year % 100:02d}.{month:02d}"


def _ocr_columns(img, v_lines, h_lines):
    """컬럼별로 "그 페이지의 모든 행을 세로로 이어붙인 이미지 1장"을 한 번만
    OCR하고, 글자의 실제 픽셀 y좌표를 우리가 이미 알고 있는 행 경계(위에서
    잘라 붙인 위치)에 직접 매핑한다. 모듈 docstring 참고 -- 셀당 OCR을 부르면
    (컬럼수 x 행수)번 프로세스가 뜨는데, 이 방식은 컬럼당 1번(페이지당 8번)
    으로 줄인다.
    """
    n_rows = len(h_lines) - 2  # 헤더 행 제외
    sep, scale = 14, 2
    columns = {}
    for c, (colname, lang, whitelist) in enumerate(COLUMN_SPECS, start=1):
        cx0, cx1 = v_lines[c] + 8, v_lines[c + 1] - 8
        strips = [
            img.crop((cx0, h_lines[r] + 6, cx1, h_lines[r + 1] - 6))
            for r in range(1, len(h_lines) - 1)
        ]
        w = max(s.width for s in strips)
        total_h = sum(s.height for s in strips) + sep * (len(strips) - 1)
        canvas = Image.new("RGB", (w, total_h), "white")
        row_bands = []
        y = 0
        for s in strips:
            canvas.paste(s, (0, y))
            row_bands.append((y, y + s.height))
            y += s.height + sep
        canvas = _darken_red_ink(canvas)
        canvas = canvas.resize((w * scale, total_h * scale), Image.LANCZOS)

        cfg = "--psm 6"
        if whitelist:
            cfg += f' -c tessedit_char_whitelist="{whitelist}"'
        data = pytesseract.image_to_data(canvas, lang=lang, config=cfg, output_type=Output.DICT)

        row_tokens = [[] for _ in range(n_rows)]
        for i, raw in enumerate(data["text"]):
            token = raw.strip()
            if not token:
                continue
            cy = (data["top"][i] + data["height"][i] / 2) / scale
            row_idx = next(
                (ri for ri, (a, b) in enumerate(row_bands) if a - sep / 2 <= cy <= b + sep / 2),
                None,
            )
            if row_idx is None:
                continue
            row_tokens[row_idx].append((data["left"][i], token))

        columns[colname] = [
            "".join(tok for _, tok in sorted(tokens, key=lambda t: t[0])) for tokens in row_tokens
        ]
    return columns


def _clean_text(raw):
    return NOISE_LEADING_RE.sub("", raw).strip()


def _parse_int(raw):
    digits = re.sub(r"[^0-9]", "", raw or "")
    return int(digits) if digits else None


def _parse_signed_int(raw):
    if raw is None:
        return None
    digits = re.sub(r"[^0-9]", "", raw)
    if not digits:
        return None
    value = int(digits)
    return -value if "-" in raw else value


def _parse_birth(raw):
    digits = re.sub(r"[^0-9]", "", raw or "")
    if len(digits) != 6:
        return None
    return f"{digits[0:2]}.{digits[2:4]}.{digits[4:6]}"


def _parse_yearmonth(raw):
    digits = re.sub(r"[^0-9]", "", raw or "")
    if len(digits) != 4:
        return None
    return f"{digits[0:2]}.{digits[2:4]}"


def extract_retirement_pdf(pdf_path, dpi=DPI, log=print):
    """스캔본 퇴직공제 PDF에서 "OO년 OO월 공제부금 납부 신고 내역" 표만 골라
    사람별 레코드 DataFrame으로 돌려준다(성명/생년월일/근로년월/근로일수/직종/
    업체명/단위공사명/납부액/needs_review/리포트월/페이지)."""
    doc = fitz.open(str(pdf_path))
    records = []
    current_report_month = None
    running_seq = 0
    pages_used = 0

    for page_index in range(doc.page_count):
        img = _render_page(doc[page_index], dpi)
        arr = np.array(img.convert("L"))
        grid = detect_table_grid(arr)
        if grid is None:
            continue
        v_lines, h_lines = grid
        pages_used += 1

        report_month = _detect_report_month(img, h_lines[0]) or current_report_month
        if report_month != current_report_month:
            current_report_month = report_month
            running_seq = 0

        columns = _ocr_columns(img, v_lines, h_lines)
        n_rows = len(h_lines) - 2
        for r in range(n_rows):
            running_seq += 1
            name = _clean_text(columns["성명"][r])
            birth = _parse_birth(columns["생년월일"][r])
            yearmonth = _parse_yearmonth(columns["근로년월"][r])
            days = _parse_signed_int(columns["근로일수"][r])
            job = _clean_text(columns["직종"][r])
            company = _clean_text(columns["업체명"][r])
            unit = _clean_text(columns["단위공사명"][r])
            amount = _parse_signed_int(columns["납부액"][r])

            sign_mismatch = (
                days is not None and amount is not None and days != 0 and amount != 0
                and (days < 0) != (amount < 0)
            )
            needs_review = (
                not name
                or birth is None
                or yearmonth is None
                or days is None
                or amount is None
                or sign_mismatch
            )
            records.append(
                {
                    "순번": running_seq,
                    "성명": name,
                    "생년월일": birth,
                    "근로년월": yearmonth,
                    "근로일수": days,
                    "직종": job,
                    "업체명": company,
                    "단위공사명": unit,
                    "납부액": amount,
                    "리포트월": current_report_month,
                    "페이지": page_index,
                    "needs_review": needs_review,
                }
            )
        log(
            f"      페이지 {page_index}: {current_report_month or '월 미확인'} "
            f"{n_rows}행 추출 (누적 {pages_used}페이지)",
        )

    doc.close()
    return pd.DataFrame(records)


def load_retirement_excel(excel_path):
    """월별 시트('26.02' 등)를 읽어 PDF 추출 결과와 같은 컬럼 이름으로 맞춘
    DataFrame 딕셔너리를 돌려준다."""
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    month_re = re.compile(r"^\d{2}\.\d{2}$")
    sheets = {}
    for sheet_name in wb.sheetnames:
        if not month_re.match(sheet_name):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = next(
            (i for i, row in enumerate(rows) if row and row[0] == "NO"), None
        )
        if header_idx is None:
            continue
        records = []
        for row in rows[header_idx + 1 :]:
            no, name, birth, yearmonth, days, job, company, unit, amount = row[0:9]
            if not isinstance(no, (int, float)) or not name:
                continue
            ym = None
            if hasattr(yearmonth, "year"):
                ym = f"{yearmonth.year % 100:02d}.{yearmonth.month:02d}"
            elif yearmonth is not None:
                digits = re.sub(r"[^0-9]", "", str(yearmonth))
                if len(digits) == 4:
                    ym = f"{digits[0:2]}.{digits[2:4]}"
            records.append(
                {
                    "순번": int(no),
                    "성명": str(name).strip(),
                    "생년월일": re.sub(r"\s+", "", str(birth)) if birth else None,
                    "근로년월": ym,
                    "근로일수_엑셀": days,
                    "직종_엑셀": job,
                    "업체명_엑셀": str(company).strip() if company else None,
                    "단위공사명_엑셀": unit,
                    "납부액_엑셀": amount,
                }
            )
        sheets[sheet_name] = pd.DataFrame(records)
    return sheets


def compare_retirement(pdf_df, excel_sheets, log=print):
    """리포트월별로 PDF OCR 결과와 엑셀 시트를 (성명, 생년월일, 근로년월,
    단위공사명) 키로 매칭해 판정한다.

    단위공사명까지 키에 넣는 이유: 실측 결과 같은 사람이 같은 달에 두 개
    이상의 단위공사(예: 배관공사/주기기설치공사)에 걸쳐 공제부금이 나뉘어
    잡히는 경우가 있어서(국민연금 쪽과 같은 문제, NOTES.md 참고), (성명,
    생년월일, 근로년월)만 키로 쓰면 이런 사람의 여러 행이 서로 겹쳐 매칭돼
    행 수가 부풀었다(실측: 818/819행짜리 두 표를 매칭했는데 976행이 나옴 --
    outer merge가 중복 키를 카티전 곱으로 늘린 것). 업체명 대신 단위공사명을
    쓰는 이유는 업체명(긴 한글 상호)보다 단위공사명(짧은 영문+숫자 코드,
    예: "MC-08")이 이 문서에서 OCR 오류가 훨씬 적었다(실측: 35행 전부
    정확). 업체명 자체는 그래도 결과 컬럼으로 남겨 육안 확인 대상이 되게
    한다(드물게 "(주)"->"(수)" 같은 오독이 있었음, NOTES.md 참고)."""
    if pdf_df.empty:
        return pd.DataFrame()

    all_results = []
    for report_month, pdf_month_df in pdf_df.groupby("리포트월"):
        excel_df = excel_sheets.get(report_month)
        if excel_df is None:
            log(f"      경고: 엑셀에 '{report_month}' 시트가 없어 대조 불가 ({len(pdf_month_df)}행)")
            continue

        merged = pd.merge(
            excel_df,
            pdf_month_df.drop(columns=["순번", "리포트월"]),
            left_on=["성명", "생년월일", "근로년월", "단위공사명_엑셀"],
            right_on=["성명", "생년월일", "근로년월", "단위공사명"],
            how="outer",
            suffixes=("", "_pdf"),
        )

        def _judge(row):
            has_excel = pd.notna(row["근로일수_엑셀"])
            has_pdf = pd.notna(row["근로일수"])
            if not has_excel and has_pdf:
                return "증빙만_존재(엑셀누락)"
            if has_excel and not has_pdf:
                return "엑셀만_존재(증빙누락)"
            if row.get("needs_review"):
                return "확인필요(OCR노이즈)"
            e_days = pd.to_numeric(row["근로일수_엑셀"], errors="coerce")
            p_days = pd.to_numeric(row["근로일수"], errors="coerce")
            e_amt = pd.to_numeric(row["납부액_엑셀"], errors="coerce")
            p_amt = pd.to_numeric(row["납부액"], errors="coerce")
            if pd.isna(e_days) or pd.isna(p_days) or pd.isna(e_amt) or pd.isna(p_amt):
                return "확인필요(OCR노이즈)"
            days_ok = e_days == p_days
            amount_ok = e_amt == p_amt
            if days_ok and amount_ok:
                return "일치"
            if p_amt > e_amt:
                return "과오납_의심(증빙>엑셀)"
            if p_amt < e_amt:
                return "과소납_의심(엑셀>증빙)"
            return "불일치"

        merged["판정"] = merged.apply(_judge, axis=1)
        merged.insert(0, "리포트월", report_month)
        all_results.append(merged)
        log(f"      [{report_month}] 엑셀 {len(excel_df)}행 / PDF {len(pdf_month_df)}행 대조 완료")

    return pd.concat(all_results, ignore_index=True) if all_results else pd.DataFrame()


def run_retirement(excel_path, pdf_path, result_dir=None, log=print):
    log(f"[retirement] PDF OCR 추출 시작: {Path(pdf_path).name} (스캔본 -- 시간이 걸립니다)")
    pdf_df = extract_retirement_pdf(pdf_path, log=log)
    log(f"[retirement] PDF에서 {len(pdf_df)}행 추출 완료")

    excel_sheets = load_retirement_excel(excel_path)
    log(f"[retirement] 엑셀 시트 {len(excel_sheets)}개 로드: {sorted(excel_sheets.keys())}")

    result = compare_retirement(pdf_df, excel_sheets, log=log)
    if result.empty:
        log("[retirement] 대조 결과 없음")
        return result

    review_ratio = result["needs_review"].fillna(False).mean()
    log(f"[retirement] 총 {len(result)}행, needs_review 비율 {review_ratio:.1%}")
    log(result["판정"].value_counts().to_string())

    if result_dir:
        result_dir = Path(result_dir)
        result_dir.mkdir(parents=True, exist_ok=True)
        out_path = result_dir / "퇴직공제_비교결과.xlsx"
        result.to_excel(out_path, index=False)
        log(f"[retirement] 저장 완료: {out_path}")

    return result


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="퇴직공제부금 OCR 대조 (스캔본 전용)")
    parser.add_argument("--excel", required=True)
    parser.add_argument("--pdf", required=True)
    parser.add_argument("--result-dir", default="result")
    args = parser.parse_args()
    run_retirement(args.excel, args.pdf, args.result_dir)
