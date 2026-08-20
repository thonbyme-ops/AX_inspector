"""건강장기요양보험료 대조 파이프라인 v4 (이슈 #2, #3) -- 네이티브 텍스트 PDF 검증.

v3와 코드는 동일하다(로직 변경 없음). v4는 "raw/대아이앤씨 건강보험 납부확인서
26.04.pdf"(공단에서 직접 발급받은 네이티브 텍스트 PDF, 스캔/OCR 아님)로
v3의 좌표·판별 로직이 그대로 통하는지 검증하려고 별도 파일로 분리했다.

검증 결과(1페이지 실측):
- 페이지 크기 595.28x841.89 (v3가 캘리브레이션한 Acrobat OCR PDF의
  598.56x843.84와 거의 동일) -- COLUMN_X_RANGES를 그대로 써도 모든 컬럼의
  실제 단어 x좌표가 기존 범위 안에 들어온다. 좌표 재조정 불필요.
- _is_format_b_text("입자부담"/"용자부담")도 정상 매치.
- 단, "네이티브 텍스트니까 노이즈가 없을 것"이라는 예상과 달리, 실제로는
  다수의 행에서 문자가 깨져 나온다(예: "16,700원"이 "_--1-6``, 700원"으로,
  이름/숫자 사이에 "J r기' 1, •" 같은 알 수 없는 토큰이 섞임) -- 아마도
  보이지 않는 워터마크/보안 텍스트 레이어가 실제 데이터와 겹쳐서 같이
  추출되는 것으로 보인다. 즉 "네이티브 = 깨끗함"이 아니었다 -- OCR발
  노이즈는 없지만 다른 종류의 텍스트 오염이 있다. needs_review 메커니즘이
  이런 행을 실제로 잡아내는지가 이번 검증의 핵심이다.

v2와의 차이(v3에서 그대로 승계):

1. OCR을 Tesseract로 다시 하지 않는다. v3는 입력 PDF가 이미 Acrobat GUI의
   "텍스트 인식"으로 OCR돼 있다고 전제한다(raw -> ocr_done은 단순 복사).
   실측 결과 Acrobat OCR 텍스트 품질이 Tesseract보다 훨씬 좋아서(예: "순번"이
   음절 단위로 안 쪼개지고 한 토큰으로 나옴), 크롭 이미지를 다시 Tesseract로
   OCR할 필요가 없어졌다 -- pdfplumber가 이미 인식해 둔 단어를 좌표로 걸러서
   바로 쓴다. 그래서 pymupdf 렌더링/PIL 크롭/pytesseract 호출이 전부 없다.

2. v2는 "PDF 한 파일 = 표 하나"를 전제했다. 실제 349페이지 번들은 회사·월별로
   문서가 반복되고, 심지어 같은 회사·월에 대해 서로 다른 두 문서 포맷이 섞여
   있다: (A) 회사 자체 제출용 "납부 내역서"(레이아웃이 지저분하고 OCR도 깨짐),
   (B) 국민건강보험공단이 발급한 공식 "납부확인서"(가입자부담/사용자부담 2줄
   구조 -- v2가 이미 캘리브레이션한 포맷과 동일). 349페이지 전수 조사 결과
   표지·기성고 개요·FORMAT A까지 섞여 있었다 -- v3는 FORMAT B만 추출 대상으로
   삼는다(사용자 확정: 공단 발급 공식 문서가 대조 기준이어야 한다).

3. 엑셀은 418개 시트(회사x월x보험종류)로 구성돼 있고, PDF 번들은 건강보험만
   담고 있다. 시트명 "{회사약칭}_건강(YY.MM)"에서 회사약칭+연월을 뽑아 PDF
   안의 해당 문서 그룹과 자동으로 짝짓는다(회사약칭은 PDF 상 회사명 텍스트의
   부분 문자열이라는 걸 418개 시트 전체에서 확인함).

norm/clean_amount/step2_extract_excel/step4_compare/step5_save_result 등
회사 단위와 무관한 로직은 전부 extract_compare_v2에서 그대로 가져다 쓴다.
"""
import argparse
import re
import shutil
import statistics
import time
from pathlib import Path

import pandas as pd
import pdfplumber

from extract_compare_v2 import (
    _find_header_idx,
    _find_single_file,
    _first_amount,
    _first_grouped_amount,
    clean_amount,
    step4_compare,
    step5_save_result,
)

BASE_DIR = Path(__file__).resolve().parent
RAW_DIR = BASE_DIR / "raw"
OCR_DONE_DIR = BASE_DIR / "ocr_done"
RESULT_DIR = BASE_DIR / "result"


def _stage_ocr_done(raw_dir, ocr_done_dir):
    """v3는 OCR을 하지 않는다 -- raw의 PDF를 ocr_done으로 그대로 복사만 한다
    (raw의 PDF가 이미 Acrobat GUI OCR로 텍스트 레이어를 갖고 있다고 전제)."""
    raw_dir = Path(raw_dir)
    ocr_done_dir = Path(ocr_done_dir)
    ocr_done_dir.mkdir(parents=True, exist_ok=True)
    for pdf_path in raw_dir.glob("*.pdf"):
        dest = ocr_done_dir / pdf_path.name
        if not dest.exists():
            shutil.copy2(pdf_path, dest)


# ---------------------------------------------------------------------------
# 페이지 분류
# ---------------------------------------------------------------------------


def _is_format_b_text(text):
    """공단 발급 공식 "납부확인서"(가입자부담/사용자부담 2줄 구조) 페이지인지 판별.

    OCR 노이즈로 "가입자부담"이 "?f입자부담"/"7R입자부담" 등으로 깨져도
    "입자부담" 부분은 349페이지 전수 조사에서 항상 살아남았다 -- 이 부분
    문자열만 확인한다("사용자부담"도 마찬가지로 "용자부담"만 확인).
    """
    return ("입자부담" in text) and ("용자부담" in text)


MONTH_RE = re.compile(r"(\d{4})년\s*(\d{1,2})월")


def _normalize_month(text):
    """텍스트에서 첫 "YYYY년 MM월"을 찾아 "YY.MM" 형식으로 돌려준다. 없으면 None."""
    m = MONTH_RE.search(text)
    if not m:
        return None
    year, month = int(m.group(1)), int(m.group(2))
    return f"{year % 100:02d}.{int(month):02d}"


ISSUE_NO_RE = re.compile(r"발\s*급\s*번?\s*호\s*:?\s*([A-Za-z0-9\-‐-―一]{5,})")


def _extract_company_text(text):
    """헤더 페이지에서 "사업장명칭" 주변 텍스트를 회사명 후보로 뽑는다.

    이 문서는 표 헤더 바로 앞줄에 회사명이 나온다(예: "거명이앤씨(주)／일용/
    공주가스발전소"). "사업장명칭" 레이블이 있는 줄 근처를 통째로 후보 텍스트로
    돌려주고, 매칭은 호출부에서 부분 문자열로 한다(정확한 셀 경계를 몰라도
    되게 하려는 의도 -- OCR 레이아웃이 페이지마다 살짝씩 다르다).
    """
    return text


def _group_format_b_pages(pdf):
    """FORMAT B 페이지를 순서대로 훑어 "문서 그룹"(회사x월 하나)으로 묶는다.

    이 문서는 연속 페이지 관례가 두 가지 섞여 있다: (1) 매 페이지마다 같은
    발급번호를 반복 인쇄하는 유형(예: 거명이앤씨 HRSG 2025.12 문서가 13~19번
    페이지에 걸쳐 있는데 각 페이지가 전부 "발급번호: EDl36-...9691273"을
    반복함), (2) 첫 페이지에만 발급번호가 있고 이후 페이지는 "1/2"->"2/2"처럼
    이어지는 유형. 그래서 "헤더가 있으면 무조건 새 그룹"으로 하면 (1) 유형이
    페이지마다 쪼개져 버린다(실측으로 확인함 -- 7페이지짜리 문서가 7개 그룹으로
    쪼개지면서 매칭 시 첫 페이지 사람들만 남고 나머지가 통째로 사라졌다).

    그래서 발급번호가 파싱되면 "직전 그룹과 발급번호가 같은가"로 같은 문서인지
    판단한다(다르면 새 그룹). 발급번호가 없는 페이지(유형 2의 연속 페이지)는
    직전 그룹에 그냥 이어붙인다. 직전 페이지가 FORMAT B가 아니었는데
    발급번호도 못 찾으면(헤더 파싱 실패) needs_review를 단 채로 새 그룹을
    만든다 -- 조용히 버리지 않는다.

    반환: [{"company_text":str, "month":str|None, "issue_no":str|None,
            "page_indices":[int,...], "needs_review":bool}, ...]
    """
    groups = []
    prev_was_format_b = False
    for i, page in enumerate(pdf.pages):
        text = page.extract_text() or ""
        if not _is_format_b_text(text):
            prev_was_format_b = False
            continue

        issue_match = ISSUE_NO_RE.search(text)
        issue_no = issue_match.group(1) if issue_match else None
        # 비교는 숫자만 남겨서 한다 -- 같은 발급번호라도 페이지마다 대시 문자가
        # "-"/"–"/"—"/"一"로 제각각 인식돼(실측 확인) 문자열 그대로 비교하면
        # 같은 문서가 다른 발급번호로 보여 그룹이 쪼개진다.
        issue_digits = re.sub(r"\D", "", issue_no) if issue_no else None
        has_header = issue_no is not None
        # 실제 발급번호는 17~19자리인데, 페이지에 따라 OCR이 심하게 잘려서
        # "36"/"3613" 같은 몇 자리만 남는 경우가 실측으로 확인됐다(대시 문자
        # 대응만으로는 못 고침). 이런 저신뢰 짧은 값은 새 문서의 시작으로 믿지
        # 않고 직전 그룹의 연속 페이지로 취급한다 -- 진짜 새 문서라면 다음
        # 페이지에서 온전한 발급번호가 나와 결국 올바르게 분리된다.
        RELIABLE_DIGITS_MIN_LEN = 10
        issue_reliable = issue_digits is not None and len(issue_digits) >= RELIABLE_DIGITS_MIN_LEN

        same_doc_as_prev = (
            prev_was_format_b
            and groups
            and (
                (not issue_reliable)  # 헤더 없음 또는 저신뢰 -- 직전 그룹에 이어붙임
                or (issue_digits == groups[-1]["issue_digits"])
            )
        )
        if same_doc_as_prev:
            groups[-1]["page_indices"].append(i)
        else:
            groups.append(
                {
                    "company_text": _extract_company_text(text) if has_header else "",
                    "month": _normalize_month(text) if has_header else None,
                    "issue_no": issue_no,
                    "issue_digits": issue_digits,
                    "page_indices": [i],
                    "needs_review": not has_header,  # 헤더 파싱 실패 안전판
                }
            )
        prev_was_format_b = True
    return groups


# ---------------------------------------------------------------------------
# 좌표 기반 추출 (Acrobat OCR 텍스트 레이어를 좌표로 필터링 -- 이미지 OCR 없음)
# ---------------------------------------------------------------------------

# Acrobat OCR PDF 실측 좌표 기준(페이지 크기 약 598x844pt) -- v2의 dpi=72
# Tesseract 렌더링 좌표계와는 스케일이 다르다(약 1/3~1/4).
COLUMN_X_RANGES = {
    "순번_pdf": (50, 95),
    "건강보험료_고지": (230, 310),
    "장기요양보험료_고지": (300, 390),
    "건강보험료_납부": (395, 465),
    "장기요양보험료_납부": (465, 545),
}
HEADER_WORD = "순번"  # 이제 음절로 안 쪼개지고 한 토큰으로 나온다 -- 정확히 매치
FALLBACK_HEADER_WORD = "생년월일"  # "순번"이 이 페이지에서만 유독 깨지는 경우가
# 실측으로 확인됐다(예: 배가건설 26.03월 문서 -- "순번" 자리에 "t ^" 같은 잡음만
# 인식됨). 이러면 top이 0으로 폴백돼 표 위 로고/워터마크 영역까지 본문으로
# 오인해서 블록을 아예 하나도 못 뽑는 문제가 있었다("순번" 한 줄 아래 있는
# "생년월일"은 여러 페이지에서 안정적으로 깨끗하게 인식됨 -- 보조 앵커로 씀).
TABLE_BODY_TOP_MARGIN = 20
FALLBACK_TABLE_BODY_TOP_MARGIN = 14  # "생년월일"은 "순번"보다 7pt 아래 있어 그만큼 줄임
TABLE_BODY_BOTTOM_MARGIN = 10
LINE_CLUSTER_GAP = 10  # 사람 내부 줄 간격(5~7pt)과 사람 간 간격(~14pt) 사이 -- 실측 보정
BLOCK_PAD_ABOVE = 8

# "사용목적"이 이제 한 토큰으로 깨끗하게 인식된다(v2 Tesseract는 음절 단위로
# 쪼개져서 후보에서 제외했었다) -- 가장 구체적이라 1순위, 나머지는 폴백.
FOOTER_MARKER_TEXTS = ("사용목적", "납부확인용", "합계", "고지", "인원")
MAX_BLOCKS_NO_FOOTER = 60

MIN_PLAUSIBLE_SUBROW_GAP = 3
MAX_PLAUSIBLE_SUBROW_GAP = 20
DEFAULT_SUBROW_HEIGHT = 13
MAX_CONSECUTIVE_EMPTY_BLOCKS = 2
SEQ_EXTRA_PAD = 3


def _cluster_lines(words, gap=LINE_CLUSTER_GAP):
    ws = sorted(words, key=lambda w: w["top"])
    clusters = []
    for w in ws:
        if clusters and w["top"] - clusters[-1][-1]["top"] <= gap:
            clusters[-1].append(w)
        else:
            clusters.append([w])
    return clusters


def _find_table_body_bounds(words, page_height):
    header_word = next((w for w in words if w["text"] == HEADER_WORD), None)
    if header_word is not None:
        top = header_word["top"] + TABLE_BODY_TOP_MARGIN
    else:
        fallback_word = next((w for w in words if w["text"] == FALLBACK_HEADER_WORD), None)
        top = (fallback_word["top"] + FALLBACK_TABLE_BODY_TOP_MARGIN) if fallback_word else 0

    footer_candidates = [
        w
        for w in words
        if w["top"] > top and any(marker in w["text"] for marker in FOOTER_MARKER_TEXTS)
    ]
    footer_word = min(footer_candidates, key=lambda w: w["top"], default=None)

    if footer_word is not None:
        return top, footer_word["top"] - TABLE_BODY_BOTTOM_MARGIN, True
    return top, page_height, False


def _typical_subrow_height(clusters):
    tops = sorted(min(w["top"] for w in c) for c in clusters)
    gaps = [b - a for a, b in zip(tops, tops[1:])]
    plausible = [g for g in gaps if MIN_PLAUSIBLE_SUBROW_GAP <= g <= MAX_PLAUSIBLE_SUBROW_GAP]
    return statistics.median(plausible) if plausible else DEFAULT_SUBROW_HEIGHT


def _words_in_box(words, x0, x1, y0, y1):
    return [w for w in words if x0 <= w["x0"] < x1 and y0 <= w["top"] < y1]


def _first_amount_v3(text):
    """v2의 _first_grouped_amount(쉼표 3자리 그룹만 매치)를 그대로 쓰되, "0원"
    (쉼표가 없어 원래 정규식이 못 잡는 값)만 예외로 0을 인식한다.

    v2가 쉼표 없는 짧은 숫자를 전부 받아들이려다 실패한 이유(문자 화이트리스트
    OCR이 "원" 같은 글자를 숫자 잡음으로 오인식해 큰 금액이 깨졌음)는 여기선
    해당 안 된다 -- v3는 크롭 OCR이 아니라 Acrobat이 이미 인식해 둔 "0원"이라는
    깨끗한 단어 토큰을 그대로 읽는 것이라, "0"이라는 토큰 자체가 통째로
    다른 문자와 섞여 있을 위험이 없다. 그래서 "0" 또는 "0원" 토큰만 좁게
    예외 처리해도 안전하다(실측: 이 문서에서 0원인 사람이 흔해서 이걸 놓치면
    멀쩡한 사람이 통째로 needs_review 없이 사라져 버렸다).
    """
    grouped = _first_grouped_amount(text)
    if grouped is not None:
        return grouped
    for token in text.split():
        if token.rstrip("원") == "0":
            return 0
    return None


# 실측(page.chars) 결과: 정상 데이터는 전부 font='LKBAPO+Dotum' size=9.0로
# 일관되는데, 오염된 글자들은 'Helvetica'/'Times-Roman'/'HiddenHorzOCR'로
# 나온다 -- 특히 "HiddenHorzOCR"이라는 이름 자체가 이 "네이티브" PDF에도
# 보이는 진짜 텍스트 아래 품질 나쁜 숨은 OCR 텍스트 레이어가 겹쳐서 깔려
# 있다는 뜻이다. 이 폰트들은 파일마다 달라질 수 있는 서브셋 폰트명(예:
# "LKBAPO+")이 아니라 PDF 표준 범용 폴백 폰트라 이름으로 안전하게 골라낼 수
# 있다 -- 이 목록으로 문자 단위로 걸러내고 단어를 다시 추출한다.
NOISE_FONTS = {"Helvetica", "Times-Roman", "Courier", "Symbol", "ZapfDingbats"}


def _is_real_content_char(obj):
    if obj.get("object_type") != "char":
        return True
    fontname = obj.get("fontname", "")
    if "Hidden" in fontname:
        return False
    if fontname in NOISE_FONTS:
        return False
    return True


def _extract_group_records(pdf, page_indices, source_label):
    """한 문서 그룹(같은 회사x월, 페이지 1개 이상)에서 사람별 레코드를 뽑는다.

    v2의 _extract_pdf_by_coordinates와 같은 알고리즘(동적 줄 클러스터링 ->
    블록 경계 계산 -> 블록별 순번/금액 파싱)이되, 크롭+Tesseract 대신 이미
    인식된 단어 텍스트를 좌표로 걸러서 바로 쓴다. 숨은 OCR 노이즈 레이어가
    있는 문서(v4에서 실측 발견)를 위해 단어 추출 전에 문자 단위로
    _is_real_content_char로 걸러낸다 -- 노이즈가 없는 문서(v3의 Acrobat OCR
    PDF 등)에서는 모든 문자가 통과하므로 동작이 그대로 유지된다.
    """
    records = []
    prev_seq = None
    for page_index in page_indices:
        page = pdf.pages[page_index].filter(_is_real_content_char)
        words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
        if not words:
            continue

        body_top, body_bottom, footer_confident = _find_table_body_bounds(words, page.height)
        body_words = [w for w in words if body_top <= w["top"] <= body_bottom]
        if not body_words:
            continue

        clusters = _cluster_lines(body_words)
        if not clusters:
            continue

        block_height = 2 * _typical_subrow_height(clusters)
        cluster_tops = sorted(min(w["top"] for w in c) for c in clusters)

        block_starts = [cluster_tops[0]]
        y = cluster_tops[0]
        while True:
            next_tops = [t for t in cluster_tops if t > y + block_height * 0.8]
            if not next_tops:
                break
            y = next_tops[0]
            block_starts.append(y)
        block_starts.append(body_bottom)

        consecutive_empty = 0
        block_count = 0
        for i in range(len(block_starts) - 1):
            if block_starts[i] >= body_bottom or consecutive_empty >= MAX_CONSECUTIVE_EMPTY_BLOCKS:
                break
            if not footer_confident and block_count >= MAX_BLOCKS_NO_FOOTER:
                break
            block_count += 1

            y0 = block_starts[i] - BLOCK_PAD_ABOVE
            y1 = block_starts[i + 1]

            seq_words = _words_in_box(
                words, *COLUMN_X_RANGES["순번_pdf"], y0 - SEQ_EXTRA_PAD, y1 + SEQ_EXTRA_PAD
            )
            seq_text = " ".join(w["text"] for w in seq_words)
            seq_no = _first_amount(seq_text)

            values = {}
            for key in ("건강보험료_고지", "장기요양보험료_고지", "건강보험료_납부", "장기요양보험료_납부"):
                cell_words = _words_in_box(words, *COLUMN_X_RANGES[key], y0, y1)
                cell_text = " ".join(w["text"] for w in cell_words)
                values[key] = _first_amount_v3(cell_text)

            all_values_empty = all(v is None for v in values.values())
            if seq_no is None and all_values_empty:
                consecutive_empty += 1
                continue
            consecutive_empty = 0
            if seq_no is not None and all_values_empty:
                continue  # 요약/합계 섹션에서 새어든 순번 잡음으로 추정 -- 기록 안 함

            # 이 서식은 "납부"(가운데 줄, 합계)가 항상 "고지"(각 절반)의 정확히
            # 2배다 -- 실측 결과 Acrobat OCR이 드물게 큰 금액의 앞자리 숫자를
            # 통째로 빠뜨리는 경우가 있었는데(예: "1,108,680원"->",108,680원",
            # 100만원 차이), 이 내부 일관성이 깨지면 인식 오류로 보고
            # needs_review로 잡는다 -- 조용히 틀린 값을 확정하지 않는다.
            amount_consistent = True
            for jigi_key, nabu_key in (
                ("건강보험료_고지", "건강보험료_납부"),
                ("장기요양보험료_고지", "장기요양보험료_납부"),
            ):
                jigi, nabu = values[jigi_key], values[nabu_key]
                if jigi is not None and nabu is not None and nabu != 2 * jigi:
                    amount_consistent = False

            needs_review = (
                seq_no is None
                or any(v is None for v in values.values())
                or (prev_seq is not None and seq_no != prev_seq + 1)
                or not amount_consistent
            )
            if seq_no is not None:
                prev_seq = seq_no
            records.append(
                {
                    "순번_pdf": seq_no,
                    "순번_후보": None,  # v3는 직접 텍스트라 다수결이 필요 없음
                    "성명": None,
                    **values,
                    "needs_review": needs_review,
                    "출처파일": source_label,
                }
            )
    return records


# ---------------------------------------------------------------------------
# 전체 시트 순회 + 회사·월 자동 매칭
# ---------------------------------------------------------------------------


PDF_RECORD_COLUMNS = [
    "순번_pdf",
    "순번_후보",
    "성명",
    "건강보험료_고지",
    "장기요양보험료_고지",
    "건강보험료_납부",
    "장기요양보험료_납부",
    "needs_review",
    "출처파일",
]


def _dedupe_by_seq(df):
    if df.empty:
        return df
    value_cols = ["건강보험료_고지", "장기요양보험료_고지", "건강보험료_납부", "장기요양보험료_납부"]
    with_seq = df[df["순번_pdf"].notna()].copy()
    without_seq = df[df["순번_pdf"].isna()]
    if not with_seq.empty:
        with_seq["_missing"] = with_seq[value_cols].isna().sum(axis=1)
        with_seq = with_seq.sort_values("_missing").drop_duplicates("순번_pdf", keep="first")
        with_seq = with_seq.drop(columns="_missing")
    return pd.concat([with_seq, without_seq], ignore_index=True)


SHEET_RE = re.compile(r"^(.+?)_건강\((\d{2}\.\d{2})\)$")


PAREN_SUFFIX_RE = re.compile(r"\([^)]*\)\s*$")


def _find_matching_groups(company_abbrev, month, groups):
    """회사약칭이 그룹의 회사명 텍스트에 부분 문자열로 포함되고 월이 일치하는
    그룹을 전부 찾아 돌려준다.

    발급번호 OCR이 페이지마다 심하게 깨지는 극소수 케이스(예: 같은 문서인데
    한 페이지만 숫자가 하나 더/덜 인식돼 그룹이 우연히 갈라짐)가 실측으로
    확인됐다 -- 그룹 단계에서 전부 못 합쳐도, 같은 회사x월로 매칭되는 그룹이
    여럿이면 여기서 레코드를 합쳐 실질적으로 복구한다. 진짜 같은 회사에
    같은 달 문서가 두 번 발급된 경우도 이론상 있을 수 있어 needs_review로
    표시는 남긴다.

    회사약칭에 "대아(주기기,배관)"/"대아(탱크)"처럼 괄호로 세부 공종을 구분한
    접미사가 붙는 경우가 있는데(v4 검증 중 실측 확인), 이 괄호 부분은 PDF
    문서의 회사명 텍스트에는 전혀 나오지 않는다(PDF는 그냥 "대아이앤씨(주)..."
    라고만 함) -- 그대로 부분 문자열 매칭을 하면 전부 매칭 실패한다. 괄호
    접미사를 뗀 기본 이름으로 먼저 시도하고, 그래도 못 찾으면 원래 약칭
    그대로도 시도한다(괄호 없는 회사는 기존 동작 그대로 유지).
    """
    base_abbrev = PAREN_SUFFIX_RE.sub("", company_abbrev).strip()
    for abbrev in (base_abbrev, company_abbrev):
        matches = [g for g in groups if g["month"] == month and abbrev in g["company_text"]]
        if matches:
            return matches
    return []


def _extract_excel_sheet_from_wb(wb, sheet_name):
    """step2_extract_excel과 같은 로직(순번/성명/건강보험료/장기요양보험료
    파싱)이지만, 이미 열려 있는 워크북 객체를 받아서 쓴다 -- step2_extract_excel은
    호출할 때마다 load_workbook으로 418개 시트짜리 파일을 통째로 다시 여는데,
    시트 200개 이상을 순회하는 v3에서는 이게 실측으로 병목이었다(시트 1개당
    수 초 이상)."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = _find_header_idx(rows)
    if header_idx is None:
        raise ValueError(f"'순번/성명' 헤더 행을 찾을 수 없습니다: (시트: {sheet_name})")

    records = []
    for row in rows[header_idx + 1:]:
        seq_no, name = row[0], row[1]
        if not isinstance(seq_no, (int, float)) or not name:
            break
        records.append(
            {
                "순번_엑셀": int(seq_no),
                "성명": str(name).strip(),
                "건강보험료_엑셀": clean_amount(row[2]),
                "장기요양보험료_엑셀": clean_amount(row[3]),
            }
        )
    return pd.DataFrame(records)


def run(excel_path, pdf_path, result_dir, sheet_filter=None):
    print(f"[1/3] PDF 문서 그룹 빌드 중: {pdf_path.name}", flush=True)
    t0 = time.time()
    with pdfplumber.open(pdf_path) as pdf:
        groups = _group_format_b_pages(pdf)
        print(f"      FORMAT B 문서 그룹 {len(groups)}개 발견 ({time.time() - t0:.1f}초)", flush=True)

        for g in groups:
            g["records"] = _extract_group_records(pdf, g["page_indices"], pdf_path.name)
    print(f"[1/3] 완료\n", flush=True)

    print("[2/3] 엑셀 시트별 매칭 + 대조 중...", flush=True)
    t0 = time.time()
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, data_only=True, read_only=True)
    sheet_names = [s for s in wb.sheetnames if "_건강(" in s]
    if sheet_filter:
        sheet_names = [s for s in sheet_names if sheet_filter in s]
    print(f"      대상 시트 {len(sheet_names)}개 (전체 {len(wb.sheetnames)}개 중 '_건강(' 포함)", flush=True)

    all_results = []
    matched_count = 0
    unmatched_sheets = []
    for sheet_name in sheet_names:
        m = SHEET_RE.match(sheet_name)
        if not m:
            continue
        company_abbrev, month = m.group(1), m.group(2)

        matched_groups = _find_matching_groups(company_abbrev, month, groups)
        if not matched_groups:
            unmatched_sheets.append(sheet_name)
            continue
        matched_count += 1

        excel_df = _extract_excel_sheet_from_wb(wb, sheet_name)
        # 그룹이 여러 개 매칭돼도(파편화 병합) 레코드 자체는 이미 각자
        # needs_review를 갖고 있다 -- 여기서 시트 전체를 일괄로 needs_review
        # 처리하면(예전엔 그렇게 했음) 실측 결과 값이 정확히 일치하는 행까지
        # "확인필요"로 덮어써 버려서 신호가 무의미해졌다. 헤더 파싱 자체가
        # 실패한 그룹(orphan)의 레코드만 needs_review를 강제한다.
        all_records = []
        for g in matched_groups:
            for r in g["records"]:
                if g["needs_review"]:
                    r = {**r, "needs_review": True}
                all_records.append(r)
        pdf_df = pd.DataFrame(all_records, columns=PDF_RECORD_COLUMNS)
        pdf_df = _dedupe_by_seq(pdf_df) if not pdf_df.empty else pdf_df

        result = step4_compare(excel_df, pdf_df)
        result.insert(0, "시트명", sheet_name)
        all_results.append(result)

    print(
        f"[2/3] 완료: {matched_count}개 시트 매칭, {len(unmatched_sheets)}개 시트 PDF_문서없음 "
        f"({time.time() - t0:.1f}초)\n",
        flush=True,
    )
    if unmatched_sheets:
        print(f"      PDF 매칭 실패 시트: {unmatched_sheets}", flush=True)

    print("[3/3] 결과 저장 중...", flush=True)
    t0 = time.time()
    combined = pd.concat(all_results, ignore_index=True) if all_results else pd.DataFrame()
    paths = step5_save_result(combined, result_dir)
    print(f"[3/3] 저장 완료 ({time.time() - t0:.1f}초): {paths}", flush=True)

    if not combined.empty:
        review_ratio = combined["needs_review"].fillna(False).mean()
        print(f"\n요약: 총 {len(combined)}행, needs_review 비율 {review_ratio:.1%}", flush=True)
        print(combined["판정"].value_counts().to_string(), flush=True)

    return combined


def _parse_args():
    parser = argparse.ArgumentParser(description="건강장기요양보험료 대조 파이프라인 v3")
    parser.add_argument("--excel", default=None)
    parser.add_argument("--pdf", default=None)
    parser.add_argument("--raw-dir", default=str(RAW_DIR))
    parser.add_argument("--ocr-done-dir", default=str(OCR_DONE_DIR))
    parser.add_argument("--result-dir", default=str(RESULT_DIR))
    parser.add_argument(
        "--sheet-filter",
        default=None,
        help="디버깅용: 시트명에 이 문자열이 포함된 시트만 처리(예: 거명_건강)",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    raw_dir = Path(args.raw_dir)
    ocr_done_dir = Path(args.ocr_done_dir)
    result_dir = Path(args.result_dir)

    excel_path = Path(args.excel) if args.excel else _find_single_file(raw_dir, "*.xlsx", "엑셀")
    if args.pdf:
        pdf_path = ocr_done_dir / args.pdf
    else:
        pdf_path = _find_single_file(ocr_done_dir, "*.pdf", "PDF(ocr_done)")

    run(excel_path, pdf_path, result_dir, sheet_filter=args.sheet_filter)
    print("\n완료! result 폴더 확인하세요.")
