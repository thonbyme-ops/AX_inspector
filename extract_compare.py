"""보험료 대조 파이프라인 (이슈 #2, #3) -- 통합 버전 (v3 + v4 + 국민연금 신규 추가).

기존 v3(Acrobat OCR 스캔본)와 v4(네이티브 PDF)는 로직이 거의 동일하고 상수만
달랐다. 이 파일은 그 공통 엔진(회사/월 자동 매칭, 엑셀 읽기, 비교, 결과 저장,
숨은 폰트 필터링)을 하나로 합치고, 문서 종류별로 다른 부분(표 판별 키워드,
컬럼 좌표, 사람당 줄 수, 엑셀 컬럼 매핑)만 `DOCUMENT_TYPES` 설정으로 뽑아냈다.

실행:
    python extract_compare.py --doc-type health --sheet-filter 대아
    python extract_compare.py --doc-type pension --pdf "대아이앤씨 국민연금 납부확인서 26.04.pdf"
    python extract_compare.py --doc-type retirement       # 스캔본 OCR 전용, retirement_ocr.py로 위임
    python extract_compare.py --doc-type all              # 문서 종류 전부 순서대로 처리 + 결과 합침

퇴직공제부금(retirement)은 이 파일의 공용 엔진(DOCUMENT_TYPES)에 안 들어있다
-- 유일한 증빙 PDF가 텍스트 레이어 없는 순수 스캔본이고 표 구조도 완전히
달라서(자세한 내용은 retirement_ocr.py 모듈 docstring, NOTES.md "퇴직공제부금
OCR 파이프라인 완성" 절 참고) 별도 OCR 모듈로 분리했다. run()에서 doc_type이
"retirement"일 때만 그쪽으로 위임한다.

숨은 폰트 필터링에 대해 (NOTES.md 참고, 2026-08-14 발견):
같은 "Hidden" 계열 폰트라도 문서 종류에 따라 정반대 의미다 -- 스캔본(Acrobat
OCR)은 Hidden 텍스트가 유일한 본문(78%)이라 걸러내면 안 되고, 네이티브 PDF는
Hidden이 눈에 보이는 진짜 텍스트 위에 겹친 중복 오염 레이어라 걸러내야 한다.
그래서 필터 적용 여부도 문서 종류 설정(`apply_noise_font_filter`)으로 분기한다
-- 스캔본 계열 문서 종류를 추가할 땐 반드시 False로 둘 것.

국민연금 관련 알려진 한계: 국민연금 엑셀 시트("_연금(YY.MM)")는 건강보험
시트와 달리 2행짜리 그룹 헤더에, 심지어 회사가 여러 하위 공사(예: 대아이앤씨의
"플랜트배관공사"/"주기기설치공사")를 진행하면 "사용자부담금" 열 자체가
공사별로 나뉘어 있다(사람마다 그중 한 열에만 값이 있고 나머지는 빈칸). 이
파일의 `_extract_excel_pension`은 그 여러 "사용자부담금" 열을 라벨로 찾아
비어있지 않은 값을 취하는 방식으로 대응했지만, 대아 시트 하나로만 검증했다
-- 다른 회사 시트에서는 레이아웃이 다를 수 있으니 실제 비교 전 반드시
샘플 검증할 것.
"""
import argparse
import re
import statistics
import time
from pathlib import Path

import numpy as np
import pandas as pd
import pdfplumber
import pytesseract
from openpyxl import load_workbook
from PIL import Image
from pytesseract import Output

from retirement_ocr import _detect_lines as _ocr_detect_grid_lines

import retirement_ocr
from extract_compare_v2 import (
    _find_header_idx,
    _find_single_file,
    _first_amount,
    _first_grouped_amount,
    clean_amount,
    step5_save_result,
)

BASE_DIR = Path(__file__).resolve().parent
RAW_DIR = BASE_DIR / "raw"
OCR_DONE_DIR = BASE_DIR / "ocr_done"
RESULT_DIR = BASE_DIR / "result"


# ---------------------------------------------------------------------------
# 공통: 숨은/오염 폰트 필터링 (NOTES.md 참고 -- 스캔본에는 절대 적용 금지)
# ---------------------------------------------------------------------------

NOISE_FONTS = {"Helvetica", "Times-Roman", "Courier", "Symbol", "ZapfDingbats"}


def _is_real_content_char(obj):
    """실측(page.chars) 결과: 정상 데이터는 서브셋 임베드 폰트(예: 'LKBAPO+Dotum')로
    일관되게 나오고, 오염된 글자들은 'Helvetica'/'Times-Roman'/'HiddenHorzOCR' 같은
    PDF 표준 범용 폴백 폰트나 이름에 "Hidden"이 들어간 폰트로 나온다. 이 폰트들은
    파일마다 달라질 수 있는 서브셋 폰트명이 아니라 표준 폰트라 이름으로 안전하게
    골라낼 수 있다.

    스캔본(Acrobat OCR)에서는 이 필터를 쓰면 안 된다 -- 그 문서는 Hidden 폰트가
    유일한 본문이라 필터링하면 실제 데이터 대부분이 사라진다(NOTES.md 실측:
    78%). 적용 여부는 문서 종류가 아니라 `_detect_needs_noise_filter`로 PDF
    마다 실제 폰트 구성을 보고 자동으로 정한다(아래 참고) -- 필터가 필요
    없다고 판단된 PDF에는 이 함수를 호출하지 않는다.
    """
    if obj.get("object_type") != "char":
        return True
    fontname = obj.get("fontname", "")
    if "Hidden" in fontname:
        return False
    if fontname in NOISE_FONTS:
        return False
    return True


def _detect_needs_noise_filter(pdf, sample_pages=3):
    """페이지 몇 개의 page.chars 폰트 분포를 보고 Hidden/노이즈 폰트를
    걸러내야 하는지 PDF마다 자동으로 판단한다(NOTES.md 체크리스트의 "1. 폰트
    분포부터 찍어본다"를 코드로 자동화).

    이걸 "문서 종류"(health/pension) 단위 고정값으로 두면 안 된다는 걸 실측
    으로 확인했다 -- 같은 "health" 종류라도 Acrobat OCR 스캔본은 Hidden이
    유일한 본문이고, 네이티브 발급 PDF는 Hidden이 진짜 텍스트 위에 겹친
    중복 오염층이라 정반대 처리가 필요하다(문서 종류가 아니라 "이 PDF가
    어떻게 만들어졌는가"에 달린 문제). 그래서 doc_type 설정이 아니라 실제
    로드된 PDF 각각에 대해 이 함수로 판단한다.

    실측: 스캔본은 서브셋 임베드 폰트("XXXXXX+폰트명" 형태)가 전혀 없이
    Hidden 폰트만 있었고(embedded_other 비율 0%), 네이티브 오염 문서는
    Hidden과 별개로 진짜 임베드 폰트가 이미 압도적 비중(88%)이었다 -- 10%를
    기준으로 삼으면 두 경우 다 안전하게 갈린다.
    """
    hidden_count = 0
    embedded_other_count = 0
    total = 0
    for page in pdf.pages[:sample_pages]:
        for c in page.chars:
            total += 1
            fontname = c.get("fontname", "")
            if "Hidden" in fontname:
                hidden_count += 1
            elif "+" in fontname:
                embedded_other_count += 1
    if total == 0 or hidden_count == 0:
        return False  # Hidden 폰트 자체가 없으면 필터링할 게 없음
    return (embedded_other_count / total) > 0.10


# ---------------------------------------------------------------------------
# 공통: 텍스트 레이어 없는 스캔본 페이지의 OCR 폴백
# ---------------------------------------------------------------------------
#
# 회사마다 같은 "건강·장기요양보험료 보험료 납부확인서" 서식인데도 일부(실측:
# 신한ACT, 광영건설, 지크아이디)는 공단 EDI 사이트에서 스캔/재출력된 이미지
# PDF로 제출돼 텍스트 레이어가 없다(전 페이지 chars=0). 좌표 기반 컬럼
# 배치(HEALTH_CONFIG/PENSION_CONFIG의 column_x_ranges)는 네이티브 PDF와
# 스캔본이 시각적으로 동일한 서식이면 그대로 재사용할 수 있으므로, 이 두
# 함수는 pdfplumber의 `page.extract_text()`/`page.extract_words()`를
# "글자가 없을 때만" OCR 결과로 대체해 나머지 파이프라인(_group_target_pages,
# _extract_group_records, _cluster_lines, _words_in_box 등)은 이 페이지가
# 네이티브인지 스캔본인지 몰라도 되게 한다 -- 새 회사별 파서를 또 만들 필요가
# 없다.
#
# 금액 오독 안전망: 이 방식은 OCR 숫자를 그대로 믿는다(퇴직공제처럼 컬럼별
# 배치 OCR로 자릿수를 다듬지 않음). 대신 HEALTH_CONFIG에 이미 있던
# `amount_consistency_pairs`(고지 x2 == 납부) 내부 일관성 체크가 원래
# "Acrobat OCR 스캔본의 자릿수 누락"을 잡으려고 만들어진 안전망이라(NOTES.md
# 참고), 이 폴백에서 생기는 오독도 같은 메커니즘으로 needs_review에 걸린다.
def _ocr_page_text(page, dpi=150):
    im = page.to_image(resolution=dpi).original
    return pytesseract.image_to_string(im, lang="kor+eng", config="--psm 6")


def _ocr_month_header_text(page, dpi=300):
    """스캔본의 "OOOO년 OO월 건강·장기요양보험료 납부내역" 제목 줄은 페이지
    전체를 한 번에 OCR하면 유독 이 줄만 깨진다(실측: 광영건설 -- "2026년
    04월"이 "20264 048"로 읽힘, 낮은/높은 dpi 둘 다 동일). 반면 그 줄만 잘라
    다시 OCR하면 정확히 읽힌다(실측 확인) -- 전체 페이지 레이아웃(위쪽의
    큼직한 로고/체크박스 제목과의 폰트 크기 차이) 때문에 tesseract가 이 줄의
    글자 높이를 헷갈리는 것으로 보인다. "내역"이라는 단어의 좌표를 먼저 찾아
    그 줄 주변만 잘라 재시도한다."""
    im = page.to_image(resolution=dpi).original
    data = pytesseract.image_to_data(im, lang="kor+eng", config="--psm 6", output_type=Output.DICT)
    anchor_top = None
    for i, raw in enumerate(data["text"]):
        if raw.strip() in ("내역", "납부내역") and data["top"][i] < im.height * 0.4:
            anchor_top = data["top"][i]
            break
    if anchor_top is None:
        return ""
    band = im.crop((0, max(0, anchor_top - 40), im.width, anchor_top + 40))
    return pytesseract.image_to_string(band, lang="kor+eng", config="--psm 6")


def _ocr_words_for_page(page, dpi=300):
    im = page.to_image(resolution=dpi).original
    scale = 72.0 / dpi  # 이미지 픽셀 -> PDF 포인트(나머지 파이프라인의 좌표계)
    data = pytesseract.image_to_data(im, lang="kor+eng", config="--psm 4", output_type=Output.DICT)
    words = []
    for i, raw in enumerate(data["text"]):
        text = raw.strip()
        if not text:
            continue
        words.append(
            {
                "text": text,
                "x0": data["left"][i] * scale,
                "top": data["top"][i] * scale,
            }
        )
    return words


# ---------------------------------------------------------------------------
# 공통: 발급번호/월 파싱, 페이지 그룹핑 (문서 종류 무관)
# ---------------------------------------------------------------------------

MONTH_RE = re.compile(r"(\d{4})년\s*(\d{1,2})월")
ISSUE_NO_RE = re.compile(r"발\s*급\s*번?\s*호\s*:?\s*([A-Za-z0-9\-‐-―一]{5,})")


def _normalize_month(text):
    m = MONTH_RE.search(text)
    if not m:
        return None
    year, month = int(m.group(1)), int(m.group(2))
    return f"{year % 100:02d}.{int(month):02d}"


def _group_target_pages(pdf, cfg):
    """문서 종류의 `is_target_page`를 통과하는 페이지를 순서대로 훑어
    "문서 그룹"(회사x월 하나)으로 묶는다 -- v3의 _group_format_b_pages와 동일한
    알고리즘이되 페이지 판별 조건만 설정에서 받는다.

    같은 문서인데 페이지마다 발급번호를 반복 인쇄하는 경우(실측 확인)와,
    첫 페이지에만 있고 이후 페이지는 이어지는 경우가 섞여 있어서 발급번호는
    "숫자만 남긴 값"으로 비교하고(대시 문자가 페이지마다 "-"/"–"/"—"/"一"로
    제각각 인식됨), 짧게 잘린 저신뢰 값은 새 문서로 믿지 않고 직전 그룹에
    이어붙인다.
    """
    groups = []
    prev_was_target = False
    for i, page in enumerate(pdf.pages):
        text = page.extract_text() or ""
        scanned = not text.strip() and len(page.chars) == 0
        if scanned:
            text = _ocr_page_text(page)  # 텍스트 레이어 없는 스캔본 폴백
        if not cfg["is_target_page"](text):
            prev_was_target = False
            continue

        issue_match = ISSUE_NO_RE.search(text)
        issue_no = issue_match.group(1) if issue_match else None
        issue_digits = re.sub(r"\D", "", issue_no) if issue_no else None
        has_header = issue_no is not None
        issue_reliable = issue_digits is not None and len(issue_digits) >= 10

        month = _normalize_month(text) if has_header else None
        if scanned and has_header:
            # 스캔본은 제목 줄("OOOO년 OO월 ...")만 유독 전체 페이지 OCR에서
            # 깨지는데(실측: _ocr_month_header_text docstring 참고), 안
            # 깨지고 "월"까지 우연히 맞아떨어지는 다른 날짜(예: 페이지 하단
            # 신청일 "2026년 06월 04일")를 잘못 집어오기도 한다(실측:
            # 광영건설 -- 청구월 04월인데 신청일 06월을 집음). 그래서 스캔본은
            # 전체 페이지 정규식 결과를 신뢰하지 않고, 제목 줄만 다시 잘라
            # OCR한 값을 우선한다.
            month = _normalize_month(_ocr_month_header_text(page)) or month

        same_doc_as_prev = (
            prev_was_target
            and groups
            and ((not issue_reliable) or (issue_digits == groups[-1]["issue_digits"]))
        )
        if same_doc_as_prev:
            groups[-1]["page_indices"].append(i)
        else:
            groups.append(
                {
                    "company_text": text if has_header else "",
                    "month": month,
                    "issue_no": issue_no,
                    "issue_digits": issue_digits,
                    "page_indices": [i],
                    "needs_review": not has_header,
                }
            )
        prev_was_target = True
    return groups


# ---------------------------------------------------------------------------
# 공통: 좌표 기반 블록 추출 (컬럼/줄 구조는 doc_type 설정에서 받음)
# ---------------------------------------------------------------------------


def _cluster_lines(words, gap):
    ws = sorted(words, key=lambda w: w["top"])
    clusters = []
    for w in ws:
        if clusters and w["top"] - clusters[-1][-1]["top"] <= gap:
            clusters[-1].append(w)
        else:
            clusters.append([w])
    return clusters


def _find_table_body_bounds(words, page_height, cfg):
    header_word = next((w for w in words if w["text"] == cfg["header_word"]), None)
    if header_word is None and len(cfg["header_word"]) > 1:
        # 일부 협력업체 PDF는 "순번" 헤더 칸이 좁아서 "순"/"번"이 한 단어가
        # 아니라 두 줄로 쪼개져 나온다(실측: 배가건설) -- 그러면 정확히
        # "순번"과 일치하는 단어가 없어 body_top이 0으로 떨어지고, 표 위쪽
        # 요약란의 "고지"(고지보험료 등)가 footer로 오인돼 본문이 통째로
        # 안 잡힌다. 첫 글자만이라도 찾아 폴백한다.
        header_word = next((w for w in words if w["text"] == cfg["header_word"][0]), None)
    top = (header_word["top"] + cfg["header_top_margin"]) if header_word else 0

    footer_markers = cfg.get("footer_markers") or ()
    footer_candidates = [
        w for w in words if w["top"] > top and any(m in w["text"] for m in footer_markers)
    ]
    footer_word = min(footer_candidates, key=lambda w: w["top"], default=None)

    if footer_word is not None:
        return top, footer_word["top"] - cfg["footer_bottom_margin"], True
    return top, page_height, False


def _typical_row_gap(clusters, cfg):
    tops = sorted(min(w["top"] for w in c) for c in clusters)
    gaps = [b - a for a, b in zip(tops, tops[1:])]
    lo, hi = cfg["min_plausible_subrow_gap"], cfg["max_plausible_subrow_gap"]
    plausible = [g for g in gaps if lo <= g <= hi]
    return statistics.median(plausible) if plausible else cfg["default_subrow_height"]


def _first_amount_safe(text):
    """v2의 _first_grouped_amount(쉼표 3자리 그룹만 매치)에 "0원"(쉼표 없어
    원래 정규식이 못 잡는 값) 예외만 추가한다. 크롭 OCR이 아니라 이미 인식된
    깨끗한 단어 텍스트를 읽는 것이라 "0"/"0원" 토큰만 좁게 허용해도 안전하다
    (v4에서 실측: 이걸 안 하면 0원인 사람이 needs_review 없이 통째로 사라짐)."""
    grouped = _first_grouped_amount(text)
    if grouped is not None:
        return grouped
    for token in text.split():
        if token.rstrip("원") == "0":
            return 0
    return None


def _words_in_box(words, x0, x1, y0, y1):
    return [w for w in words if x0 <= w["x0"] < x1 and y0 <= w["top"] < y1]


def _ocr_cell_amount(im, x0_pt, x1_pt, y0_pt, y1_pt, scale, signed=True):
    """포인트 좌표 박스 하나를 잘라 숫자 전용으로 OCR한다(퇴직공제 OCR과
    같은 방식: 셀을 통째로 넓게 잡고 word-clustering에 맡기는 대신, 이미 알고
    있는 칸 하나만 딱 잘라 psm 7(한 줄)로 읽으면 훨씬 정확하다 -- 실측:
    워터마크가 겹친 행에서도 이 방식은 살아남음).

    signed=False(순번처럼 항상 양수인 컬럼)일 땐 화이트리스트에서 "-"를
    아예 뺀다 -- 안 그러면 셀 오른쪽 격자선이 살짝 걸려서 "-"로 오독되는
    경우가 있다(실측: 순번 "1"이 "1-"로 읽힘 -> 파싱 실패)."""
    pad = 3  # 포인트 -- 격자선이 살짝 걸리면 psm 7이 아예 텍스트를 못 찾는다(실측)
    box = (
        int((x0_pt + pad) * scale),
        int((y0_pt + pad) * scale),
        int((x1_pt - pad) * scale),
        int((y1_pt - pad) * scale),
    )
    if box[2] <= box[0] or box[3] <= box[1]:
        return None
    crop = im.crop(box)
    crop = crop.resize((crop.width * 2, crop.height * 2), Image.LANCZOS)
    whitelist = "-0123456789," if signed else "0123456789"
    text = pytesseract.image_to_string(
        crop, lang="eng", config=f'--psm 7 -c tessedit_char_whitelist="{whitelist}"'
    ).strip()
    if not signed:
        # 순번처럼 쉼표 없는 작은 수는 _first_grouped_amount(쉼표 3자리 그룹만
        # 매치)로 못 잡는다 -- _first_amount(그룹 여부 무관하게 첫 숫자 토큰)
        # 를 쓴다.
        return _first_amount(text)
    grouped = _first_grouped_amount(text)
    if grouped is not None:
        return -grouped if "-" in text else grouped
    for token in text.split():
        if token.rstrip("원").lstrip("-") == "0":
            return 0
    return None


def _ocr_extract_group_records_grid(page, cfg, source_label, prev_seq=None, dpi=300):
    """스캔본 전용 대체 추출기. 실측(광영건설): 워터마크/음영이 겹친 행에서는
    `page.extract_words()`/`_ocr_words_for_page`의 단어 단위 OCR이 그 행만
    통째로 놓친다(전체 페이지를 한 번에 인식시키기 때문). 사람 행 경계는
    격자선(표 테두리, 검은 픽셀) 검출로 찾는다 -- 이건 워터마크에 영향받지
    않는다(실측 확인). 컬럼 좌표는 새로 검출하지 않고 같은 서식의 네이티브
    PDF로 이미 검증된 `cfg["column_x_ranges"]`를 그대로 재사용한다(시각적으로
    동일한 공단 표준 서식이라 좌표가 그대로 맞음, NOTES.md 참고). 값은 셀
    하나씩 잘라 OCR하므로(퇴직공제 OCR과 같은 방식) 단어 단위 OCR보다 느리지만
    워터마크에 강하다.
    """
    im = page.to_image(resolution=dpi).original
    arr = np.array(im.convert("L"))
    dark = arr < 190
    scale = dpi / 72.0

    words = _ocr_words_for_page(page, dpi=150)
    body_top, body_bottom, _ = _find_table_body_bounds(words, page.height, cfg)

    x_ranges = cfg["column_x_ranges"]
    table_x0 = min(r[0] for r in x_ranges.values())
    table_x1 = max(r[1] for r in x_ranges.values())
    h_lines_px = _ocr_detect_grid_lines(
        dark, "h", int(table_x0 * scale), int(table_x1 * scale), 0.5
    )
    h_lines = [y / scale for y in h_lines_px if body_top - 5 <= y / scale <= body_bottom + 5]
    if len(h_lines) < cfg["block_height_multiplier"] + 1:
        return [], prev_seq

    mult = cfg["block_height_multiplier"]
    value_columns = cfg["value_columns"]
    seq_range = x_ranges["순번_pdf"]

    records = []
    for i in range(0, len(h_lines) - mult, mult):
        y0, y1 = h_lines[i], h_lines[i + mult]
        seq_no = _ocr_cell_amount(im, *seq_range, y0, y1, scale, signed=False)
        values = {key: _ocr_cell_amount(im, *x_ranges[key], y0, y1, scale) for key in value_columns}

        if seq_no is None and all(v is None for v in values.values()):
            continue

        amount_consistent = True
        for a_key, b_key in cfg.get("amount_consistency_pairs", []):
            a, b = values.get(a_key), values.get(b_key)
            if a is not None and b is not None and b != 2 * a:
                amount_consistent = False

        needs_review = (
            seq_no is None
            or any(v is None for v in values.values())
            or (prev_seq is not None and seq_no != prev_seq + 1)
            or not amount_consistent
        )
        if seq_no is not None:
            prev_seq = seq_no
        records.append(
            {"순번_pdf": seq_no, **values, "needs_review": needs_review, "출처파일": source_label}
        )
    return records, prev_seq


def _extract_group_records(pdf, page_indices, source_label, cfg, apply_noise_filter):
    """한 문서 그룹(같은 회사x월, 페이지 1개 이상)에서 사람별 레코드를 뽑는다.

    v3/v4의 _extract_group_records와 같은 알고리즘(동적 줄 클러스터링 -> 블록
    경계 계산 -> 블록별 순번/금액 파싱)이되, 컬럼 x범위·줄 간격·블록 높이
    배수 등을 전부 doc_type 설정에서 받아 문서 종류에 상관없이 동작한다.
    "block_height_multiplier"가 그 차이를 흡수한다 -- 건강보험처럼 사람당
    여러 줄(가입자부담/사용자부담)이 겹쳐 있으면 typical 줄 간격의 2배가
    한 블록 높이지만, 국민연금처럼 사람당 정확히 한 줄이면 배수가 1이라
    typical 줄 간격이 곧 블록 높이다.

    apply_noise_filter는 doc_type 설정이 아니라 이 PDF에 대해
    _detect_needs_noise_filter가 실제로 판단한 값을 호출부(run_doc_type)가
    넘겨준다 -- 문서 종류로 고정하면 안 되는 이유는 _is_real_content_char
    docstring 참고.
    """
    x_ranges = cfg["column_x_ranges"]
    value_columns = cfg["value_columns"]

    records = []
    prev_seq = None
    for page_index in page_indices:
        page = pdf.pages[page_index]
        if len(page.chars) == 0:
            # 텍스트 레이어 없는 스캔본: 단어 단위 OCR은 워터마크 겹친 행을
            # 통째로 놓치므로(실측: 광영건설) 격자선 기반 대체 추출기를 쓴다
            # (_ocr_extract_group_records_grid docstring 참고).
            page_records, prev_seq = _ocr_extract_group_records_grid(
                page, cfg, source_label, prev_seq=prev_seq
            )
            records.extend(page_records)
            continue
        if apply_noise_filter:
            page = page.filter(_is_real_content_char)
        words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
        if not words:
            continue

        body_top, body_bottom, footer_confident = _find_table_body_bounds(
            words, page.height, cfg
        )
        body_words = [w for w in words if body_top <= w["top"] <= body_bottom]
        if not body_words:
            continue

        clusters = _cluster_lines(body_words, cfg["line_cluster_gap"])
        if not clusters:
            continue

        block_height = cfg["block_height_multiplier"] * _typical_row_gap(clusters, cfg)
        cluster_tops = sorted(min(w["top"] for w in c) for c in clusters)

        block_starts = [cluster_tops[0]]
        y = cluster_tops[0]
        while True:
            next_tops = [t for t in cluster_tops if t > y + block_height * 0.8]
            if not next_tops:
                break
            y = next_tops[0]
            block_starts.append(y)
        block_starts.append(body_bottom)

        consecutive_empty = 0
        block_count = 0
        max_blocks_no_footer = cfg["max_blocks_no_footer"]
        for i in range(len(block_starts) - 1):
            if block_starts[i] >= body_bottom or consecutive_empty >= cfg["max_consecutive_empty_blocks"]:
                break
            if not footer_confident and block_count >= max_blocks_no_footer:
                break
            block_count += 1

            y0 = block_starts[i] - cfg["block_pad_above"]
            y1 = block_starts[i + 1]

            seq_words = _words_in_box(
                words, *x_ranges["순번_pdf"], y0 - cfg["seq_extra_pad"], y1 + cfg["seq_extra_pad"]
            )
            seq_text = " ".join(w["text"] for w in seq_words)
            seq_no = _first_amount(seq_text)

            values = {}
            for key in value_columns:
                cell_words = _words_in_box(words, *x_ranges[key], y0, y1)
                cell_text = " ".join(w["text"] for w in cell_words)
                values[key] = _first_amount_safe(cell_text)

            all_values_empty = all(v is None for v in values.values())
            if seq_no is None and all_values_empty:
                consecutive_empty += 1
                continue
            consecutive_empty = 0
            if seq_no is not None and all_values_empty:
                continue  # 요약/합계 섹션에서 새어든 순번 잡음으로 추정 -- 기록 안 함

            amount_consistent = True
            for a_key, b_key in cfg.get("amount_consistency_pairs", []):
                a, b = values.get(a_key), values.get(b_key)
                if a is not None and b is not None and b != 2 * a:
                    amount_consistent = False

            needs_review = (
                seq_no is None
                or any(v is None for v in values.values())
                or (prev_seq is not None and seq_no != prev_seq + 1)
                or not amount_consistent
            )
            if seq_no is not None:
                prev_seq = seq_no
            records.append(
                {
                    "순번_pdf": seq_no,
                    **values,
                    "needs_review": needs_review,
                    "출처파일": source_label,
                }
            )
    return records


def _dedupe_by_seq(df, value_columns):
    if df.empty:
        return df
    with_seq = df[df["순번_pdf"].notna()].copy()
    without_seq = df[df["순번_pdf"].isna()]
    if not with_seq.empty:
        with_seq["_missing"] = with_seq[value_columns].isna().sum(axis=1)
        with_seq = with_seq.sort_values("_missing").drop_duplicates("순번_pdf", keep="first")
        with_seq = with_seq.drop(columns="_missing")
    return pd.concat([with_seq, without_seq], ignore_index=True)


# ---------------------------------------------------------------------------
# 건강·장기요양보험료 문서 종류
# ---------------------------------------------------------------------------


def _is_health_target_page(text):
    """공단 발급 공식 "납부확인서"(가입자부담/사용자부담 2줄 구조) 페이지인지
    판별. OCR 노이즈로 "가입자부담"이 "?f입자부담" 등으로 깨져도 "입자부담"
    부분은 349페이지 전수 조사에서 항상 살아남았다."""
    return ("입자부담" in text) and ("용자부담" in text)


def _extract_excel_health(wb, sheet_name):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = _find_header_idx(rows)
    if header_idx is None:
        raise ValueError(f"'순번/성명' 헤더 행을 찾을 수 없습니다: (시트: {sheet_name})")
    records = []
    for row in rows[header_idx + 1 :]:
        seq_no, name = row[0], row[1]
        if not isinstance(seq_no, (int, float)) or not name:
            break
        records.append(
            {
                "순번_엑셀": int(seq_no),
                "성명": str(name).strip(),
                "건강보험료_엑셀": clean_amount(row[2]),
                "장기요양보험료_엑셀": clean_amount(row[3]),
            }
        )
    return pd.DataFrame(records)


HEALTH_CONFIG = {
    "label": "건강·장기요양보험료",
    "sheet_marker": "_건강(",
    "is_target_page": _is_health_target_page,
    "header_word": "순번",
    "header_top_margin": 20,
    "footer_bottom_margin": 10,
    # "사용목적"이 Acrobat OCR에서 한 토큰으로 깨끗하게 인식돼(예전 Tesseract는
    # 음절 단위로 쪼갰음) 1순위 마커로 쓴다. 나머지는 폴백.
    "footer_markers": ("사용목적", "납부확인용", "합계", "고지", "인원"),
    "max_blocks_no_footer": 60,
    "line_cluster_gap": 10,
    "block_pad_above": 8,
    "seq_extra_pad": 3,
    "block_height_multiplier": 2,  # 가입자부담/사용자부담 두 줄이 한 블록
    "min_plausible_subrow_gap": 3,
    "max_plausible_subrow_gap": 20,
    "default_subrow_height": 13,
    "max_consecutive_empty_blocks": 2,
    "column_x_ranges": {
        "순번_pdf": (50, 95),
        "건강보험료_고지": (230, 310),
        "장기요양보험료_고지": (300, 390),
        "건강보험료_납부": (395, 465),
        "장기요양보험료_납부": (465, 545),
    },
    "value_columns": ["건강보험료_고지", "장기요양보험료_고지", "건강보험료_납부", "장기요양보험료_납부"],
    # "납부"(가운데 줄, 합계)는 항상 "고지"(각 절반)의 정확히 2배다 -- 실측
    # 결과 Acrobat/네이티브 OCR이 드물게 큰 금액의 앞자리를 통째로 빠뜨리는
    # 경우가 있었는데(예: "1,108,680원"->",108,680원"), 이 내부 일관성이
    # 깨지면 needs_review로 잡는다.
    "amount_consistency_pairs": [
        ("건강보험료_고지", "건강보험료_납부"),
        ("장기요양보험료_고지", "장기요양보험료_납부"),
    ],
    "extract_excel": _extract_excel_health,
    "compare_pairs": [
        ("건강보험료_엑셀", "건강보험료_납부", "건강보험료"),
        ("장기요양보험료_엑셀", "장기요양보험료_납부", "장기요양보험료"),
    ],
    # 숨은 폰트 필터 적용 여부는 doc_type 고정값이 아니라 PDF마다
    # _detect_needs_noise_filter로 자동 판단한다(run_doc_type 참고) --
    # 이 종류에 스캔본과 네이티브 PDF가 둘 다 올 수 있어서다.
}


# ---------------------------------------------------------------------------
# 국민연금 문서 종류 (신규)
# ---------------------------------------------------------------------------


def _is_pension_target_page(text):
    """국민연금보험료 결정내역서의 "연금보험료 대상자" 표 페이지 판별.
    실측(대아이앤씨 26.04, 5페이지): 첫 페이지에만 발급번호/집계 섹션이
    있고 이후 페이지는 표 헤더("순번 성명 생년월일 기준소득월액 연금보험료
    근로자기여금 사용자부담금")만 반복하며 이어진다 -- "연금보험료"와
    "기준소득월액"이 둘 다 있으면 대상자 표가 있는 페이지로 본다."""
    return ("연금보험료" in text) and ("기준소득월액" in text)


def _extract_excel_pension(wb, sheet_name):
    """국민연금 엑셀 시트를 읽는다.

    건강보험 시트와 달리 2행짜리 그룹 헤더고(예: "납부보험료"가 상위 그룹,
    그 아래 실제 "사용자부담금"/"노무비 지급내역서 No." 열들이 옴), 회사가
    여러 하위 공사를 진행하면 "사용자부담금" 열 자체가 공사별로 나뉘어
    있다(대아이앤씨 26.04 실측: "플랜트배관공사"/"주기기설치공사" 각각의
    사용자부담금 열이 따로 있고, 사람마다 그중 한 열에만 값이 있음).

    그래서 "사용자부담금"이라는 라벨을 가진 열을 전부 찾아서, 각 행마다 그중
    비어있지 않은 첫 값을 취한다(한 사람이 동시에 여러 공사에서 일하지
    않는다고 전제). "연금보험료"(총액) 열은 헤더 행에 직접 라벨이 없고 한
    행 위의 그룹 헤더("납부보험료")에 있어서, 순번/성명 바로 다음(성명 뒤
    첫 번째, "사용자부담금" 라벨이 아닌) 열로 찾는다.

    주의: 대아이앤씨 시트 하나로만 검증했다 -- 다른 회사 시트는 하위 공사가
    하나뿐이라 레이아웃이 더 단순하거나, 반대로 또 다르게 생겼을 수 있다.
    실제 비교 전 반드시 대상 시트로 샘플 검증할 것.
    """
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = _find_header_idx(rows)
    if header_idx is None:
        raise ValueError(f"'순번/성명' 헤더 행을 찾을 수 없습니다: (시트: {sheet_name})")

    header_row = rows[header_idx]
    burden_cols = [i for i, v in enumerate(header_row) if isinstance(v, str) and "사용자부담금" in v]
    if not burden_cols:
        raise ValueError(f"'사용자부담금' 열을 찾을 수 없습니다: (시트: {sheet_name})")
    # "연금보험료"(총액) 열: 성명(1) 바로 다음 칸부터 "사용자부담금"이 아닌 첫 열
    pension_col = next((i for i in range(2, max(burden_cols) + 1) if i not in burden_cols), 2)

    records = []
    for row in rows[header_idx + 1 :]:
        seq_no, name = row[0], row[1]
        if not isinstance(seq_no, (int, float)) or not name:
            break
        burden = next((clean_amount(row[i]) for i in burden_cols if row[i] not in (None, "")), None)
        records.append(
            {
                "순번_엑셀": int(seq_no),
                "성명": str(name).strip(),
                "연금보험료_엑셀": clean_amount(row[pension_col]) if pension_col < len(row) else None,
                "사용자부담금_엑셀": burden,
            }
        )
    return pd.DataFrame(records)


PENSION_CONFIG = {
    "label": "국민연금",
    "sheet_marker": "_연금(",
    "is_target_page": _is_pension_target_page,
    "header_word": "순번",
    "header_top_margin": 15,
    "footer_bottom_margin": 10,
    # 실측 결과 대상자 표 페이지 자체에는 뚜렷한 마커 텍스트가 안 보였다
    # (건강보험과 달리 요약 섹션이 첫 페이지에만 따로 있음) -- 표시 못 찾으면
    # max_blocks_no_footer 폴백에 의존한다. 한 페이지 최대 인원 관측치(약
    # 24명)보다 넉넉하게 잡는다.
    "footer_markers": (),
    "max_blocks_no_footer": 40,
    "line_cluster_gap": 10,
    "block_pad_above": 3,
    "seq_extra_pad": 3,
    "block_height_multiplier": 1,  # 사람당 정확히 한 줄
    "min_plausible_subrow_gap": 15,
    "max_plausible_subrow_gap": 35,
    "default_subrow_height": 24,
    "max_consecutive_empty_blocks": 2,
    "column_x_ranges": {
        "순번_pdf": (35, 62),
        "연금보험료": (295, 365),
        "근로자기여금": (380, 460),
        "사용자부담금": (478, 560),
    },
    "value_columns": ["연금보험료", "근로자기여금", "사용자부담금"],
    "amount_consistency_pairs": [],  # 근로자기여금==사용자부담금이지만 필수 관계로 강제하지 않음
    "extract_excel": _extract_excel_pension,
    "compare_pairs": [
        ("연금보험료_엑셀", "연금보험료", "연금보험료"),
        ("사용자부담금_엑셀", "사용자부담금", "사용자부담금"),
    ],
    # 숨은 폰트 필터는 여기서도 doc_type 고정값이 아니라 PDF마다 자동 판단
    # (HEALTH_CONFIG 주석 참고).
}


DOCUMENT_TYPES = {
    "health": HEALTH_CONFIG,
    "pension": PENSION_CONFIG,
}

# 퇴직공제부금은 이 파일의 나머지 문서 종류와 근본적으로 다르게 처리한다 --
# 유일한 증빙 PDF(`sample/06-05. 제20회 기성 실적정산(퇴직공제).pdf`)가
# 텍스트 레이어 없는 순수 스캔본(전 페이지 chars=0)이라 pdfplumber로 못 읽고,
# 표 자체도 "회사x월 페이지 그룹"이 아니라 "월 1개 표에 전체 협력업체가
# 업체명 컬럼으로 뒤섞인 한 장" 구조라 이 파일의 공용 엔진(_group_target_pages
# 이하)과 맞지 않는다. 그래서 OCR 전용 모듈(retirement_ocr.py)로 완전히
# 분리했다 -- DOCUMENT_TYPES 제네릭 엔진에는 넣지 않고 run()에서 doc_type이
# "retirement"일 때만 별도 분기한다(아래 run() 참고).
RETIREMENT_DOC_TYPE = "retirement"
ALL_DOC_TYPES = [*DOCUMENT_TYPES.keys(), RETIREMENT_DOC_TYPE]


# ---------------------------------------------------------------------------
# 공통: 회사·월 자동 매칭 + 대조 + 결과 저장
# ---------------------------------------------------------------------------

SHEET_RE_TEMPLATE = r"^(.+?)_{marker}\((\d{{2}}\.\d{{2}})\)$"


def _find_matching_groups(company_abbrev, month, groups):
    return [g for g in groups if g["month"] == month and company_abbrev in g["company_text"]]


def _compare_generic(excel_df, pdf_df, compare_pairs):
    """문서 종류의 compare_pairs(엑셀컬럼, pdf컬럼, 표시이름) 목록을 기반으로
    순번 매칭 + 항목별 일치 판정을 한다 -- v2의 step4_compare을 문서 종류
    무관하게 일반화한 버전."""
    pdf_df = pdf_df.drop(columns=["성명"], errors="ignore")
    if "순번_pdf" in pdf_df.columns:
        # 스캔본 OCR 폴백(_ocr_extract_group_records_grid)은 순번을 못 읽으면
        # None을 넣는데, 그러면 컬럼 dtype이 object가 돼 엑셀 쪽(항상 int)과
        # 병합할 때 pandas가 타입 불일치로 에러를 낸다.
        pdf_df["순번_pdf"] = pd.to_numeric(pdf_df["순번_pdf"], errors="coerce")
    merged = pd.merge(excel_df, pdf_df, left_on="순번_엑셀", right_on="순번_pdf", how="outer")

    columns = ["성명", "순번_엑셀", "순번_pdf"]
    for excel_col, pdf_col, label in compare_pairs:
        diff_col = f"{label}_차이"
        merged[diff_col] = merged[excel_col] - merged[pdf_col]
        columns += [excel_col, pdf_col, diff_col]

    def _judge(row):
        # "행이 존재하는가"는 순번 자체로 판단한다 -- compare_pairs 값으로
        # 판단하면 안 되는 이유(실측: 거명이앤씨): 그 달 근무가 없어 보험료가
        # 0원인 사람을 엑셀이 빈 셀(공란)로 남겨두는 경우가 있는데, 그러면
        # PDF는 정확히 0원을 명시하는데도 엑셀 값이 없다는 이유만으로
        # "증빙만_존재"로 잘못 갈린다 -- 둘 다 "0원"이라는 같은 사실을
        # 말하고 있는데도.
        has_excel = pd.notna(row["순번_엑셀"])
        has_pdf = pd.notna(row["순번_pdf"])
        if not has_excel and has_pdf:
            return "증빙만_존재(엑셀누락)"
        if has_excel and not has_pdf:
            return "엑셀만_존재(증빙누락)"
        if row.get("needs_review"):
            return "확인필요(OCR노이즈)"
        # 일부 협력업체 엑셀은 이전 달 귀속분을 정정하는 행을 음수로 기록한다
        # (실측: 동남건설 "26.03 변경" 비고, 건강보험료=-206700원 -- 크기는
        # PDF의 정상 청구액과 정확히 같고 부호만 반대). PDF(공단 발급
        # 납부확인서)는 항상 그 달의 정상 청구액만 양수로 찍혀 나와서 이런
        # 정정 행과는 애초에 비교 대상이 아니다 -- 그대로 두면 "과오납_의심"
        # 으로 잘못 분류돼 실제로는 없는 초과청구처럼 보인다.
        def _num_or_zero(v):
            # `v or 0`을 안 쓰는 이유: 엑셀의 빈 셀은 merge 후 float('nan')이
            # 되는데, `nan or 0`은 nan을 그대로 돌려준다(nan은 파이썬에서
            # falsy가 아님) -- 그러면 이후 `nan != 0`이 항상 True가 되고
            # `nan < p_val`은 항상 False가 되어 매번 "과소납_의심"으로 잘못
            # 판정됐다(실측: 거명이앤씨, 빈 셀 29건 전부 오탐).
            return 0 if pd.isna(v) else v

        if any(_num_or_zero(row.get(excel_col)) < 0 for excel_col, _, _ in compare_pairs):
            return "확인필요(엑셀정정행)"

        all_ok = True
        has_diff = False
        overpaid = False
        underpaid = False
        for excel_col, pdf_col, _ in compare_pairs:
            e_val = _num_or_zero(row.get(excel_col))
            p_val = _num_or_zero(row.get(pdf_col))
            if e_val != p_val:
                all_ok = False
                has_diff = True
                if e_val < p_val:
                    overpaid = True  # 증빙 금액이 엑셀보다 큼 (초과청구/과오납 위험)
                else:
                    underpaid = True # 엑셀 금액이 증빙보다 큼 (납부부족/과소납 위험)

        if all_ok:
            return "일치"
        if overpaid and not underpaid:
            return "과오납_의심(증빙>엑셀)"
        if underpaid and not overpaid:
            return "과소납_의심(엑셀>증빙)"
        return "불일치"

    merged["판정"] = merged.apply(_judge, axis=1)
    columns += ["판정", "needs_review", "출처파일"]
    for col in columns:
        if col not in merged.columns:
            merged[col] = None
    return merged[columns].sort_values(["순번_엑셀", "순번_pdf"]).reset_index(drop=True)


def run_doc_type(doc_type, excel_path, pdf_path, sheet_filter=None):
    """한 문서 종류(health/pension)에 대해 PDF 그룹 빌드 -> 엑셀 시트 매칭 ->
    대조까지 전부 수행하고, 합쳐진 결과 DataFrame과 매칭 요약을 돌려준다."""
    cfg = DOCUMENT_TYPES[doc_type]
    print(f"[{doc_type}] PDF 문서 그룹 빌드 중: {pdf_path.name}", flush=True)
    t0 = time.time()
    with pdfplumber.open(pdf_path) as pdf:
        apply_noise_filter = _detect_needs_noise_filter(pdf)
        print(
            f"      숨은/노이즈 폰트 필터: {'적용' if apply_noise_filter else '미적용'} "
            f"(이 PDF의 실제 폰트 구성으로 자동 판단)",
            flush=True,
        )
        groups = _group_target_pages(pdf, cfg)
        print(f"      {cfg['label']} 문서 그룹 {len(groups)}개 발견 ({time.time() - t0:.1f}초)", flush=True)
        for g in groups:
            g["records"] = _extract_group_records(
                pdf, g["page_indices"], pdf_path.name, cfg, apply_noise_filter
            )

    wb = load_workbook(excel_path, data_only=True, read_only=True)
    sheet_names = [s for s in wb.sheetnames if cfg["sheet_marker"] in s]
    if sheet_filter:
        sheet_names = [s for s in sheet_names if sheet_filter in s]
    print(f"      대상 시트 {len(sheet_names)}개 ('{cfg['sheet_marker']}' 포함)", flush=True)

    sheet_re = re.compile(SHEET_RE_TEMPLATE.format(marker=cfg["sheet_marker"].strip("_(")))
    value_columns = cfg["value_columns"]

    all_results = []
    matched_count = 0
    unmatched_sheets = []
    for sheet_name in sheet_names:
        m = sheet_re.match(sheet_name)
        if not m:
            continue
        company_abbrev, month = m.group(1), m.group(2)
        # 국민연금 시트는 하위 공사 구분이 회사약칭 뒤에 괄호로 붙는다(예:
        # "대아(주기기,배관)_연금(26.04)") -- 그 괄호 부분은 엑셀 내부
        # 표기일 뿐 PDF 회사명 텍스트에는 나오지 않으므로, 매칭에는 괄호
        # 앞의 순수 약칭만 쓴다.
        company_key = company_abbrev.split("(")[0]

        matched_groups = _find_matching_groups(company_key, month, groups)
        if not matched_groups:
            unmatched_sheets.append(sheet_name)
            continue
        matched_count += 1

        excel_df = cfg["extract_excel"](wb, sheet_name)
        all_records = []
        for g in matched_groups:
            for r in g["records"]:
                if g["needs_review"]:
                    r = {**r, "needs_review": True}
                all_records.append(r)
        pdf_df = pd.DataFrame(all_records, columns=["순번_pdf", *value_columns, "needs_review", "출처파일"])
        pdf_df = _dedupe_by_seq(pdf_df, value_columns) if not pdf_df.empty else pdf_df

        result = _compare_generic(excel_df, pdf_df, cfg["compare_pairs"])
        result.insert(0, "문서종류", cfg["label"])
        result.insert(1, "시트명", sheet_name)
        all_results.append(result)

    print(
        f"[{doc_type}] 완료: {matched_count}개 시트 매칭, {len(unmatched_sheets)}개 시트 PDF_문서없음 "
        f"({time.time() - t0:.1f}초)",
        flush=True,
    )
    if unmatched_sheets:
        print(f"      PDF 매칭 실패 시트: {unmatched_sheets}", flush=True)

    combined = pd.concat(all_results, ignore_index=True) if all_results else pd.DataFrame()
    return combined


def run(doc_types, excel_path, pdf_paths, result_dir, sheet_filter=None, retirement_excel_path=None):
    """doc_types 목록(예: ["health"] 또는 ["health","pension","retirement"])을
    순서대로 처리해 결과를 하나로 합쳐 저장한다. pdf_paths는 {doc_type: Path}
    매핑. retirement는 건강/연금과 엔진 자체가 달라(모듈 위쪽 주석 참고)
    retirement_ocr.run_retirement()으로 따로 처리하고, 그 전용 엑셀(퇴직공제부금
    납부 신고 내역.xlsx)도 건강/연금이 쓰는 엑셀(보험료 납부_단위공사별.xlsx)과
    달라서 별도 경로로 받는다."""
    all_combined = []
    for doc_type in doc_types:
        pdf_path = pdf_paths[doc_type]
        if doc_type == RETIREMENT_DOC_TYPE:
            if retirement_excel_path is None:
                print("경고: 퇴직공제 전용 엑셀(예: 퇴직공제부금 납부 신고 내역.xlsx)을 찾지 못해 건너뜁니다.", flush=True)
                continue
            combined = retirement_ocr.run_retirement(retirement_excel_path, pdf_path, result_dir=None, log=lambda m: print(m, flush=True))
            if not combined.empty:
                combined.insert(0, "문서종류", "퇴직공제부금")
                combined.insert(1, "시트명", combined["리포트월"])
        else:
            combined = run_doc_type(doc_type, excel_path, pdf_path, sheet_filter=sheet_filter)
        if not combined.empty:
            all_combined.append(combined)
        print(flush=True)

    final = pd.concat(all_combined, ignore_index=True) if all_combined else pd.DataFrame()
    if final.empty:
        print("경고: 매칭된 시트가 하나도 없어 저장할 결과가 없습니다 (result 파일을 만들지 않음).", flush=True)
        return final

    print("결과 저장 중...", flush=True)
    paths = step5_save_result(final, result_dir)
    print(f"저장 완료: {paths}", flush=True)

    review_ratio = final["needs_review"].fillna(False).mean()
    print(f"\n요약: 총 {final.shape[0]}행, needs_review 비율 {review_ratio:.1%}", flush=True)
    print(final.groupby("문서종류")["판정"].value_counts().to_string(), flush=True)

    return final


def _parse_args():
    parser = argparse.ArgumentParser(description="보험료 대조 파이프라인 (통합, 문서 종류별 --doc-type)")
    parser.add_argument(
        "--doc-type",
        required=True,
        choices=[*ALL_DOC_TYPES, "all"],
        help="처리할 문서 종류. 'all'이면 전부 순서대로 처리 후 결과를 합침",
    )
    parser.add_argument("--excel", default=None)
    parser.add_argument(
        "--pdf",
        default=None,
        help="단일 --doc-type일 때만 사용. 생략하면 raw 폴더에서 해당 문서 종류 PDF를 자동 탐색",
    )
    parser.add_argument("--raw-dir", default=str(RAW_DIR))
    parser.add_argument("--result-dir", default=str(RESULT_DIR))
    parser.add_argument(
        "--sheet-filter",
        default=None,
        help="디버깅용: 시트명에 이 문자열이 포함된 시트만 처리(예: 대아)",
    )
    return parser.parse_args()


def _find_default_excel(search_dirs):
    for d in search_dirs:
        p = Path(d)
        if not p.exists():
            continue
        candidates = list(p.glob("*.xlsx"))
        # 보험료 납부_단위공사별 우선 탐색
        named = [c for c in candidates if "보험료" in c.name or "단위공사" in c.name]
        if named:
            return named[0]
        if candidates:
            return candidates[0]
    return None


def _find_retirement_excel(search_dirs):
    """퇴직공제는 health/pension이 쓰는 엑셀(보험료 납부_단위공사별.xlsx)이
    아니라 별도 파일(퇴직공제부금 납부 신고 내역.xlsx)을 쓴다 -- 월별 시트
    ('26.04' 등)로만 구성돼 있어 이름으로 구분한다."""
    for d in search_dirs:
        p = Path(d)
        if not p.exists():
            continue
        candidates = [c for c in p.glob("*.xlsx") if "퇴직공제" in c.name]
        if candidates:
            return candidates[0]
    return None


def _find_pdf_for_doc_type(search_dirs, doc_type):
    """지정된 디렉토리들에서 해당 문서 종류로 보이는 PDF를 이름 키워드로 자동 탐색한다."""
    keyword_map = {
        "health": "건강",
        "pension": "국민연금",
        "retirement": "퇴직공제",
    }
    keyword = keyword_map.get(doc_type, doc_type)
    for d in search_dirs:
        p = Path(d)
        if not p.exists():
            continue
        candidates = [f for f in p.glob("*.pdf") if keyword in f.name]
        if len(candidates) >= 1:
            return candidates[0]
    raise FileNotFoundError(f"탐색 디렉토리({search_dirs})에서 '{keyword}'가 포함된 PDF를 찾을 수 없습니다.")


if __name__ == "__main__":
    args = _parse_args()
    search_dirs = [Path(args.raw_dir), BASE_DIR / "sample", BASE_DIR / "raw", BASE_DIR / "ocr_done"]
    result_dir = Path(args.result_dir)

    if args.excel:
        excel_path = Path(args.excel)
    else:
        excel_path = _find_default_excel(search_dirs)
        if not excel_path:
            raise FileNotFoundError(f"탐색 경로({search_dirs})에서 엑셀 원장(*.xlsx)을 찾을 수 없습니다.")

    doc_types = ALL_DOC_TYPES if args.doc_type == "all" else [args.doc_type]

    if args.pdf and len(doc_types) > 1:
        raise SystemExit("--pdf는 --doc-type이 단일 문서 종류일 때만 지정할 수 있습니다.")

    pdf_paths = {}
    for dt in doc_types:
        pdf_paths[dt] = Path(args.pdf) if args.pdf else _find_pdf_for_doc_type(search_dirs, dt)

    retirement_excel_path = None
    if RETIREMENT_DOC_TYPE in doc_types:
        retirement_excel_path = _find_retirement_excel(search_dirs)
        if not retirement_excel_path:
            raise FileNotFoundError(f"탐색 경로({search_dirs})에서 퇴직공제 전용 엑셀(*퇴직공제*.xlsx)을 찾을 수 없습니다.")

    print(f"[*] 엑셀 파일: {excel_path}")
    if retirement_excel_path:
        print(f"[*] 퇴직공제 전용 엑셀: {retirement_excel_path}")
    for dt, p in pdf_paths.items():
        print(f"[*] PDF 증빙 [{dt}]: {p}")

    run(doc_types, excel_path, pdf_paths, result_dir, sheet_filter=args.sheet_filter, retirement_excel_path=retirement_excel_path)
    print("\n완료! result 폴더를 확인하세요.")
