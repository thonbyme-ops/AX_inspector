"""정규화한 노무비 출역 데이터와 각종 보험료 부과 내역을 대조한다 (이슈 #5-5 부분 착수).

담당자 의견 5번은 "실 출입기록으로 각종 보험료와 비교. 단, 건강 및 요양보험은 가입
해지 여부에 따라 차월에 반영되고, 하도급사에서 공단 신고가 늦어지는 경우 한달씩
차이가 발생하므로 이를 수동 검사해야 함"이었다.

**출입기록 엑셀 샘플이 아직 없어서**, 그 자리에 이슈 #5-3에서 정규화한 노무비 명세서의
일자별 출역을 대신 넣었다(출입기록이 들어오면 같은 대조 로직에 소스만 하나 더 붙이면
된다). 노무비는 하도급사가 돈을 지급한 근거라 "실제 일한 사람" 목록으로는 이미 신뢰도가
있고, 아래 실측처럼 퇴직공제 신고와 93% 매칭된다.

## 왜 "불일치=문제"로 단정하지 않는가 (실측 근거)

퇴직공제부금은 근로일수 비례라 노무비 출역일수와 직접 맞춰볼 수 있는 유일한 비목이다.
실측(3사 5개 파일 357명 매칭) 결과 (노무비 출역일수 - 퇴직공제 근로일수) 분포:

| 업체 / 월 | 0일 | 주요 편차 |
|---|---|---|
| 신한에이씨티 26.03 | 73% | -1일 13%, -2일 5% |
| 신한에이씨티 26.04 | 82% | -1일 9%, -2일 4% |
| NSC 26.04 | 71% | -2일 12%, -1일 10% |
| **신보 26.03** | **25%** | **+1일 53%** |

신한/NSC는 0일이 71~82%인데 신보 26.03만 +1일이 절반이다. 90명 가까이 똑같이 1일씩
어긋나는 건 개인별 누락이 아니라 **그 업체·그 달의 신고 규칙 차이**일 가능성이 훨씬
높다. 그래서 이 모듈은 차이를 판정하지 않고,
(1) 업체x월별 **최빈 편차**를 구해 그 편차와 같은 행은 `업체공통편차`로 묶고,
(2) 최빈 편차에서 벗어난 행만 `일수차이_확인필요`로 올린다.
이러면 "규칙 차이"와 "개별 이상"이 섞이지 않는다.

## 차월 반영 처리

담당자가 지적한 대로 건강·장기요양은 당월에 안 잡히는 게 정상일 수 있다. 그래서 비목별로
당월 -> 차월(M+1) -> 전월(M-1) 순으로 찾아보고 어디서 잡혔는지를 그대로 표시한다
(`당월` / `차월(2026-05)` / `전월(2026-03)` / `없음`). 건강·장기요양이 인접월에서 잡히면
`차월반영_추정`으로, 아무 달에도 없으면 `미부과_확인필요`로 분류한다.
"""
import re

from openpyxl import Workbook

from hr_cost_extractor import CATEGORY_LABELS

# 같은 회사가 자료마다 다르게 적힌다. 아래 별칭은 추측이 아니라 성명 교집합으로
# 실측해 확정한 것이다(1순위 대응이 64~85%, 2순위는 1~2%로 뚜렷하게 갈렸다):
#   노무비 "(주)신한에이씨티" <-> 보험료 원장 "(주)신한ACT"   (26.03 50/50, 26.04 58/62)
#   노무비 "NSC"             <-> 보험료 원장 "(주)엔에스컴퍼니" (26.04 45/70)
COMPANY_ALIASES = {
    "신한에이씨티": "신한ACT",
    "NSC": "엔에스컴퍼니",
}
LEGAL_FORM_RE = re.compile(r"(주식회사|\(주\)|㈜|\(유\)|유한회사|\(사\))")

# 건강·장기요양은 가입/해지가 차월에 반영되고 하도급사 신고 지연도 겹쳐 한 달씩
# 어긋나는 게 정상 범위다(담당자 의견). 연금·퇴직공제는 당월 기준이라 인접월에서
# 잡히면 그대로 확인 대상으로 올린다.
LAGGING_CATEGORIES = {"health", "longterm"}
CHECK_CATEGORIES = ["health", "longterm", "pension", "retirement"]

# 국민연금은 60세 도달로 가입 자격이 끝나므로 60세 이상은 미부과가 정상이다.
# 실측으로 확인함: 이 현장 383명 중 60세 이상 71명은 68명(96%)이 연금 미부과인 반면,
# 60세 미만 226명은 36명(16%)만 미부과였다 -- 나이를 안 보면 이 68명이 전부
# "미부과 확인필요"로 올라와 정작 봐야 할 건을 덮어버린다.
PENSION_EXEMPT_AGE = 60

# 일용근로자는 한 달 8일(또는 60시간) 미만이면 건강·연금 가입 대상이 아니다.
# 실측으로 뒷받침됨: 이 현장에서 출역 4~7일인 21명 중 19명(90%)이 건강보험 미부과인
# 반면 15일 이상인 284명은 28명(10%)만 미부과였다. 퇴직공제부금은 1일부터 대상이라
# 이 임계를 적용하지 않는다(실측: 출역 4~7일도 21명 전원 부과).
#
# 단, 이 판단은 **이 현장 노무비 명세서의 출역일수**만 보고 하는 것이라 단정할 수
# 없다 -- 같은 달에 다른 현장에서도 일하면 합산되어 8일을 넘긴다. 실제로 1~3일
# 구간은 오히려 53%가 부과돼 있었다. 그래서 "적용제외_추정"으로만 내리고 근거
# 일수를 판정 문구에 남긴다.
MIN_DAYS_FOR_COVERAGE = 8
DAY_THRESHOLD_CATEGORIES = {"health", "longterm", "pension"}


def _age(birth6, year_month):
    """주민번호 앞 6자리와 귀속월로 만 나이를 구한다.

    앞 2자리만으로는 1900년대/2000년대가 모호한데, 이 자료는 근로자 명부라
    "귀속연도보다 늦게 태어난 사람은 없다"는 것만으로 가른다(2026년 기준
    '27'~'99'는 1900년대, '00'~'26'은 2000년대).
    """
    if not birth6 or len(birth6) < 6 or not birth6.isdigit():
        return None
    ref_year, ref_month = (int(p) for p in year_month.split("-"))
    yy = int(birth6[:2])
    born_year = 2000 + yy if 2000 + yy <= ref_year else 1900 + yy
    age = ref_year - born_year
    if ref_month < int(birth6[2:4]):
        age -= 1
    return age


def company_key(name):
    """자료 간 업체명 표기 차이를 흡수한 매칭 키."""
    text = LEGAL_FORM_RE.sub("", re.sub(r"\s+", "", str(name or "")))
    return COMPANY_ALIASES.get(text, text)


def _shift_month(year_month, delta):
    year, month = (int(p) for p in year_month.split("-"))
    index = year * 12 + (month - 1) + delta
    return f"{index // 12:04d}-{index % 12 + 1:02d}"


# 부과 내역이 들어올 수 있는 두 갈래. 공단이 직접 발급한 확인서/결정내역서(이슈 #5-1의
# `confirmation_extractor`)는 생년월일이 있어 동명이인을 가를 수 있고 공단 표기 합계와
# 원 단위까지 일치가 확인된 자료라 우선한다. 자체 납부 원장은 성명뿐이라 폴백.
#
# **소스를 나누는 이유는 중복 합산 방지가 먼저다**: 같은 사람·같은 달이 원장과 PDF에
# 모두 있는데 한 통에 담아 더하면 금액이 두 배가 된다(실제로 대아이앤씨 26.04는 두
# 자료에 다 들어있다). 그래서 소스별로 따로 담고 우선순위대로 하나만 채택한다.
SOURCE_EVIDENCE = "evidence"  # 공단 발급 PDF
SOURCE_LEDGER = "ledger"      # 자체 납부 원장 / 신고 내역 엑셀
SOURCE_PRIORITY = (SOURCE_EVIDENCE, SOURCE_LEDGER)
SOURCE_LABELS = {SOURCE_EVIDENCE: "공단증빙", SOURCE_LEDGER: "자체원장"}


def _premium_index(premium_records):
    """부과 내역을 (소스 x 조회범위)별로 색인한다.

    조회범위는 성명 기준과 (성명+생년월일) 기준 두 벌이다 -- 생년월일이 있으면 그걸
    쓰고 없으면 성명으로 떨어진다. 실측으로 필요성이 드러났다: NSC 26.04에 "박정규"가
    두 명(62세/48세) 있는데 성명만으로 매칭하니 둘 다 같은 퇴직공제 레코드(두 사람
    합계인 52일)에 붙어 양쪽 다 -27일 차이로 오탐이 났다.
    """
    index = {
        source: {"name": ({}, {}), "birth": ({}, {})} for source in SOURCE_PRIORITY
    }
    # 같은 성명 키에 서로 다른 생년월일이 들어오면 그 키로 매칭한 값은 두 사람의
    # 합이다. 실측: 대아이앤씨 26.04 건강보험 확인서에 성명이 같고 생년월일이 다른
    # 사람이 두 쌍 있어, 성명으로만 집으면 두 사람 금액의 합이 나왔다.
    births_per_name = {}
    for record in premium_records:
        source = SOURCE_EVIDENCE if record.get("detail") else SOURCE_LEDGER
        company = company_key(record["company"])
        base = (company, record["person"], record["year_month"], record["category"])
        targets = [("name", base)]
        birth6 = record.get("birth6")
        if birth6:
            births_per_name.setdefault(base, set()).add(birth6)
            targets.append(
                ("birth", (company, record["person"], birth6, record["year_month"], record["category"]))
            )
        for scope, key in targets:
            amounts, workdays = index[source][scope]
            amounts[key] = amounts.get(key, 0) + (record["amount"] or 0)
            days = record.get("workdays")
            if days is not None:
                workdays[key] = workdays.get(key, 0) + days

    ambiguous = {key for key, births in births_per_name.items() if len(births) > 1}
    return {"sources": index, "ambiguous_names": ambiguous}


def _keys_for(company, person, birth6, category, year_month):
    return (
        ("birth", (company, person, birth6, year_month, category) if birth6 else None),
        ("name", (company, person, year_month, category)),
    )


def _lookup(index, company, person, birth6, category, year_month):
    """당월 -> 차월 -> 전월 순으로 찾아 (표시문구, 금액, 잡힌 연월, 조회범위, 소스,
    성명모호 여부)를 돌려준다. 같은 달 안에서는 공단 증빙을 자체 원장보다 우선한다.

    금액 0원은 "부과 안 됨"이 아니라 "그 달에 0원으로 부과됨"이다(실측: 대아이앤씨
    26.04 원장에 0원 행이 있다). 그래서 키 존재 여부로 판단한다 -- 값이 0일 때
    다음 달을 뒤지면 엉뚱한 달 금액을 그 사람 것으로 붙이게 된다.
    """
    for delta, label in ((0, "당월"), (1, "차월"), (-1, "전월")):
        target = _shift_month(year_month, delta)
        for source in SOURCE_PRIORITY:
            for scope, key in _keys_for(company, person, birth6, category, target):
                if key is None or key not in index["sources"][source][scope][0]:
                    continue
                amount = index["sources"][source][scope][0][key]
                text = label if delta == 0 else f"{label}({target})"
                if amount == 0:
                    text += "(0원)"
                name_key = (company, person, target, category)
                ambiguous = scope == "name" and name_key in index["ambiguous_names"]
                return text, amount, target, scope, source, ambiguous
    return "없음", None, None, None, None, False


def _other_source_amount(index, company, person, birth6, category, year_month, source):
    """같은 사람·달을 다른 소스에서도 찾아 금액을 돌려준다(증빙 vs 원장 대조용)."""
    other = SOURCE_LEDGER if source == SOURCE_EVIDENCE else SOURCE_EVIDENCE
    for scope, key in _keys_for(company, person, birth6, category, year_month):
        if key is None:
            continue
        amounts = index["sources"][other][scope][0]
        if key in amounts:
            return amounts[key]
    return None


def cross_check(labor_summaries, premium_records, day_diff_tolerance=1.0):
    """노무비 인별 요약 x 보험료 부과 내역을 대조한다.

    day_diff_tolerance: 업체x월 최빈 편차에서 이만큼까지는 개별 이상으로 보지 않는다.
    """
    index = _premium_index(premium_records)
    rows = []

    # 같은 업체·월에 성명이 겹치는 사람 -- 생년월일 없이 성명으로만 매칭하는 비목
    # (건강·연금)에서는 서로의 값을 집어올 수 있어 판정을 신뢰할 수 없다.
    name_counts = {}
    for s in labor_summaries:
        key = (company_key(s["company"]), s["person"], s["year_month"])
        name_counts[key] = name_counts.get(key, 0) + 1

    for summary in labor_summaries:
        company = company_key(summary["company"])
        person = summary["person"]
        year_month = summary["year_month"]
        if not (company and person and year_month):
            continue

        row = {
            "company": summary["company"],
            "company_key": company,
            "person": person,
            "birth6": summary.get("birth6"),
            "job": summary.get("job"),
            "year_month": year_month,
            "출역일수": summary["출역일수_계산"],
            "총공수": summary["총공수_계산"],
            "실지급액": summary.get("실지급액"),
        }

        age = _age(summary.get("birth6"), year_month)
        row["연령"] = age

        lagging = []
        adjacent = []
        missing = []
        exempt = []
        short_days = []
        birth6 = summary.get("birth6")
        row["동명이인"] = name_counts.get((company, person, year_month), 1) > 1
        evidence_gaps = []
        ambiguous_hits = []
        for category in CHECK_CATEGORIES:
            label, amount, hit_month, scope, source, ambiguous = _lookup(
                index, company, person, birth6, category, year_month
            )
            row[category] = label
            row[f"{category}_금액"] = amount
            row[f"{category}_매칭"] = scope
            row[f"{category}_소스"] = SOURCE_LABELS.get(source)
            if ambiguous:
                ambiguous_hits.append(category)

            # 공단 증빙과 자체 원장에 같은 사람·달이 둘 다 있으면 금액을 맞춰본다 --
            # 다르면 그게 곧 "증빙 대비 원장 불일치"다.
            if hit_month and source:
                other = _other_source_amount(
                    index, company, person, birth6, category, hit_month, source
                )
                if other is not None and other != amount:
                    evidence = amount if source == SOURCE_EVIDENCE else other
                    ledger = other if source == SOURCE_EVIDENCE else amount
                    evidence_gaps.append(
                        f"{CATEGORY_LABELS[category]} 증빙 {evidence:,} vs 원장 {ledger:,}"
                    )
            if label == "없음":
                if category == "pension" and age is not None and age >= PENSION_EXEMPT_AGE:
                    row[category] = f"적용제외({age}세)"
                    exempt.append(category)
                elif (
                    category in DAY_THRESHOLD_CATEGORIES
                    and summary["출역일수_계산"] < MIN_DAYS_FOR_COVERAGE
                ):
                    row[category] = f"적용제외추정({summary['출역일수_계산']}일)"
                    short_days.append(category)
                else:
                    missing.append(category)
            elif label != "당월":
                # 인접월에서 잡힌 것은 "미부과"가 아니다 -- 건강·장기요양은 차월 반영이
                # 정상 범위(담당자 의견)이고, 연금·퇴직공제는 신고 지연 의심이라
                # 확인 대상이긴 하지만 아예 안 낸 것과는 구분해서 표시한다.
                (lagging if category in LAGGING_CATEGORIES else adjacent).append(category)
            if category == "retirement" and hit_month:
                key = (
                    (company, person, birth6, hit_month, category)
                    if scope == "birth"
                    else (company, person, hit_month, category)
                )
                row["퇴직공제_근로일수"] = index["sources"][source][scope][1].get(key)

        reported = row.get("퇴직공제_근로일수")
        row["일수차이"] = (
            round(summary["출역일수_계산"] - reported, 1) if reported is not None else None
        )
        row["_missing"] = missing
        row["_lagging"] = lagging
        row["_adjacent"] = adjacent
        row["_exempt"] = exempt
        row["_short_days"] = short_days
        row["_evidence_gaps"] = evidence_gaps
        row["_ambiguous_hits"] = ambiguous_hits
        row["증빙대조"] = "; ".join(evidence_gaps)
        # 값이 잡힌 비목이 전부 생년월일로 매칭됐는지 -- 그러면 동명이인이어도
        # 서로의 값을 집을 수 없어 판정을 믿을 수 있다.
        matched = [row[f"{c}_매칭"] for c in CHECK_CATEGORIES if row[f"{c}_매칭"]]
        row["_birth_matched"] = bool(matched) and all(s == "birth" for s in matched)
        rows.append(row)

    _classify(rows, day_diff_tolerance)
    reverse = _reverse_rows(labor_summaries, premium_records)
    return rows, reverse, _summarize(rows)


def _modal_diff(rows):
    """업체x월별 최빈 일수차이. 다수가 똑같이 어긋나면 개인 문제가 아니라
    그 업체·그 달의 신고 규칙 차이로 보고 기준선으로 쓴다(모듈 상단 표 참고)."""
    buckets = {}
    for row in rows:
        if row["일수차이"] is None:
            continue
        buckets.setdefault((row["company_key"], row["year_month"]), []).append(row["일수차이"])
    modal = {}
    for key, diffs in buckets.items():
        counts = {}
        for diff in diffs:
            counts[diff] = counts.get(diff, 0) + 1
        best = max(counts.items(), key=lambda kv: (kv[1], -abs(kv[0])))
        # 최빈값이 과반이 아니면 기준선으로 삼지 않는다(0을 기준으로 둔다).
        modal[key] = best[0] if best[1] * 2 >= len(diffs) else 0.0
    return modal


def _labels(categories):
    return ", ".join(CATEGORY_LABELS[c] for c in categories)


def _classify(rows, tolerance):
    """판정을 우선 확인(review)과 참고(note)로 나눈다 -- 섞으면 진짜 봐야 할 건이
    정상 범위의 차월 반영에 묻힌다(실측: 나이·인접월을 구분 안 했을 때 383명 중
    209명이 "확인필요"로 올라왔다)."""
    modal = _modal_diff(rows)
    for row in rows:
        baseline = modal.get((row["company_key"], row["year_month"]), 0.0)
        row["업체공통편차"] = baseline

        review, note = [], []
        if row["_missing"]:
            review.append(f"미부과_확인필요({_labels(row['_missing'])})")
        if row["_adjacent"]:
            review.append(f"인접월반영_신고지연의심({_labels(row['_adjacent'])})")

        diff = row["일수차이"]
        if diff is None:
            note.append("퇴직공제_신고없음" if not row["_missing"] else None)
        elif abs(diff - baseline) > tolerance:
            review.append(f"일수차이_확인필요({diff:+g}일, 업체공통 {baseline:+g}일)")

        if row["_evidence_gaps"]:
            review.append("증빙대비_원장불일치(" + "; ".join(row["_evidence_gaps"]) + ")")

        if row["_ambiguous_hits"]:
            # 부과 자료 쪽에 같은 성명이 여러 명 있는데 생년월일 없이 성명으로
            # 집었다 -- 그 금액은 두 사람의 합일 수 있다.
            review.append(
                f"부과자료_동명이인합산주의({_labels(row['_ambiguous_hits'])})"
            )
        elif row["동명이인"] and not row["_birth_matched"]:
            # 생년월일로 갈린 비목은 동명이인이어도 안전하다. 성명만으로 매칭된
            # 비목이 하나라도 있으면 서로의 값을 집었을 수 있어 확인 대상에 올린다.
            review.append("동명이인_수동확인")

        if row["_lagging"]:
            note.append(f"차월반영_추정({_labels(row['_lagging'])})")
        if row["_exempt"]:
            note.append(f"연금_적용제외({row['연령']}세)")
        if row["_short_days"]:
            note.append(
                f"적용제외_추정(출역 {row['출역일수']}일 < {MIN_DAYS_FOR_COVERAGE}일: "
                f"{_labels(row['_short_days'])})"
            )

        note = [n for n in note if n]
        row["판정"] = " / ".join(review + note) if (review or note) else "정상"
        row["needs_review"] = bool(review)
        for key in ("_missing", "_lagging", "_adjacent", "_exempt", "_short_days",
                    "_evidence_gaps", "_birth_matched", "_ambiguous_hits"):
            del row[key]


def _reverse_rows(labor_summaries, premium_records):
    """보험료는 부과됐는데 노무비 명세서에 없는 사람 (유령 인력 / 노무비 미제출 의심).

    노무비 명세서를 제출한 (업체, 월) 조합만 본다 -- 아예 제출 안 한 업체까지 넣으면
    현장 전체 인원이 통째로 올라와 의미가 없다.
    """
    covered = {(company_key(s["company"]), s["year_month"]) for s in labor_summaries}
    known = {(company_key(s["company"]), s["person"], s["year_month"]) for s in labor_summaries}

    grouped = {}
    for record in premium_records:
        key = (company_key(record["company"]), record["person"], record["year_month"])
        if key[0:1] + key[2:3] not in {(c, y) for c, y in covered}:
            continue
        if (key[0], key[2]) not in covered or key in known:
            continue
        entry = grouped.setdefault(
            key, {"company": record["company"], "person": record["person"], "year_month": record["year_month"]}
        )
        entry[record["category"]] = entry.get(record["category"], 0) + (record["amount"] or 0)
    return sorted(grouped.values(), key=lambda r: (r["company"] or "", r["year_month"] or "", r["person"] or ""))


def _summarize(rows):
    summary = {}
    for row in rows:
        key = (row["company"], row["year_month"])
        agg = summary.setdefault(key, {"people": 0, "정상": 0, "확인필요": 0, "차월반영_추정": 0})
        agg["people"] += 1
        if not row["needs_review"]:
            agg["정상"] += 1
        else:
            agg["확인필요"] += 1
        if "차월반영_추정" in row["판정"]:
            agg["차월반영_추정"] += 1
    return [
        {"company": company, "year_month": year_month, **agg}
        for (company, year_month), agg in sorted(summary.items(), key=lambda kv: (kv[0][0] or "", kv[0][1] or ""))
    ]


# ---------------------------------------------------------------------------
# 엑셀 내보내기
# ---------------------------------------------------------------------------

CROSS_HEADERS = [
    ("업체명", "company"),
    ("연월", "year_month"),
    ("성명", "person"),
    ("생년월일6", "birth6"),
    ("연령", "연령"),
    ("직종", "job"),
    ("출역일수", "출역일수"),
    ("총공수", "총공수"),
    ("퇴직공제 근로일수", "퇴직공제_근로일수"),
    ("일수차이", "일수차이"),
    ("업체공통편차", "업체공통편차"),
    ("건강보험", "health"),
    ("장기요양", "longterm"),
    ("국민연금", "pension"),
    ("퇴직공제", "retirement"),
    ("부과자료 출처", "health_소스"),
    ("실지급액", "실지급액"),
    ("동명이인", "동명이인"),
    ("증빙대조", "증빙대조"),
    ("판정", "판정"),
]


def build_cross_check_workbook(rows, reverse, summary):
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "출역x보험료 대조"
    sheet.append([label for label, _ in CROSS_HEADERS])
    for row in sorted(rows, key=lambda r: (r["company"] or "", r["year_month"] or "", r["person"] or "")):
        sheet.append([row.get(key) for _, key in CROSS_HEADERS])

    review = workbook.create_sheet("확인필요만")
    review.append([label for label, _ in CROSS_HEADERS])
    for row in sorted(
        (r for r in rows if r["needs_review"]),
        key=lambda r: (r["company"] or "", r["year_month"] or "", r["person"] or ""),
    ):
        review.append([row.get(key) for _, key in CROSS_HEADERS])

    ghost = workbook.create_sheet("보험료만 존재")
    ghost.append(["업체명", "연월", "성명", *[CATEGORY_LABELS[c] for c in CHECK_CATEGORIES]])
    for row in reverse:
        ghost.append([row["company"], row["year_month"], row["person"],
                      *[row.get(c) for c in CHECK_CATEGORIES]])

    stats = workbook.create_sheet("업체월별 요약")
    stats.append(["업체명", "연월", "인원", "정상", "확인필요", "차월반영 추정"])
    for row in summary:
        stats.append([row["company"], row["year_month"], row["people"],
                      row["정상"], row["확인필요"], row["차월반영_추정"]])

    return workbook
