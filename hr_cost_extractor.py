"""보험료(건강·연금)/퇴직공제부금 엑셀에서 인당·업체별·월별 납부액을 추출한다 (이슈 #2).

지원하는 두 가지 원본 양식:
1) "보험료 납부_단위공사별" 형태 — 업체별 시트가 `{업체}_건강(YY.MM)` / `{업체}_연금(YY.MM)`
   이름으로 나뉘어 있고, 시트 안에 순번/성명/납부보험료가 인당 한 행씩 있다.
2) "퇴직공제부금 납부 신고 내역" 형태 — 월별 시트(`YY.MM`)에 전체 업체가 섞여 있고,
   행마다 성명/업체명/근로년월/납부액이 있다.

두 양식 모두 텍스트 기반 엑셀이라 pdfplumber/OCR 없이 openpyxl만으로 안정적으로
파싱할 수 있다. 결과는 (업체, 성명, 연월, 구분, 금액) 형태의 평평한 레코드 목록으로
통일해 반환하며, 화면 표시/엑셀 내보내기는 이 레코드를 집계해서 만든다.
"""
import re

from openpyxl import load_workbook, Workbook

from extractor import clean_amount

CATEGORY_LABELS = {
    "health": "건강보험료",
    "pension": "연금보험료",
    "retirement": "퇴직공제부금",
}
CATEGORY_ORDER = ["health", "pension", "retirement"]

PREMIUM_SHEET_RE = re.compile(r"^(.+?)_(건강|연금)\(\s*([\d.]+)\s*\)\s*$")
RETIREMENT_SHEET_RE = re.compile(r"^\s*(\d{2})[.\-](\d{2})\s*$")

PREMIUM_CATEGORY_MAP = {"건강": "health", "연금": "pension"}

EXCEL_ERROR_VALUES = {"#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}


def norm(s):
    return re.sub(r"\s+", "", str(s or ""))


def _yy_mm_to_iso(yy_dot_mm):
    m = re.match(r"^(\d{2})[.\-](\d{2})$", yy_dot_mm.strip())
    if not m:
        return None
    yy, mm = m.groups()
    return f"20{yy}-{mm}"


def _normalize_company(name):
    """같은 회사가 문서마다 다르게 표기되는 것을 통일한다.
    (예: "배 가 건 설 (주)" / "배가건설㈜" / "배가건설(주)" 모두 같은 회사)"""
    if not name:
        return name
    s = re.sub(r"\s+", "", str(name))
    return s.replace("㈜", "(주)")


def _find_company_name(rows, max_scan=10, fallback=None):
    for row in rows[:max_scan]:
        for idx, cell in enumerate(row):
            if cell is None or "업체명" not in norm(cell):
                continue
            m = re.search(r"업\s*체\s*명\s*[:：]\s*(.*)$", str(cell))
            value = m.group(1).strip() if m else ""
            if not value:
                # 라벨과 값이 서로 다른 셀에 나뉜 경우 (예: "업 체 명 : " | "배가건설(주)")
                for next_cell in row[idx + 1:]:
                    if next_cell is not None and str(next_cell).strip():
                        value = str(next_cell).strip()
                        break
            if value and value not in EXCEL_ERROR_VALUES:
                return _normalize_company(value)
    return _normalize_company(fallback)


def _find_header_row(rows, keyword="성명", max_scan=10):
    for idx, row in enumerate(rows[:max_scan]):
        for cell in row:
            if cell is not None and keyword in norm(cell):
                return idx
    return None


def _parse_premium_sheet(ws, category, year_month):
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    company = _find_company_name(rows, fallback=ws.title.split("_", 1)[0])
    header_idx = _find_header_row(rows)
    if header_idx is None:
        return []

    records = []
    for row in rows[header_idx + 1:]:
        seq, name = row[0] if len(row) > 0 else None, row[1] if len(row) > 1 else None
        if seq is None and not name:
            break
        if not isinstance(seq, (int, float)) or not name:
            continue
        amount = clean_amount(row[2]) if len(row) > 2 else None
        if amount is None:
            continue
        records.append(
            {
                "company": company,
                "person": str(name).strip(),
                "year_month": year_month,
                "category": category,
                "amount": amount,
            }
        )
    return records


def _parse_retirement_sheet(ws, sheet_year_month):
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    header_idx = _find_header_row(rows)
    if header_idx is None:
        return []

    records = []
    for row in rows[header_idx + 1:]:
        no, name = row[0] if len(row) > 0 else None, row[1] if len(row) > 1 else None
        if no is None and not name:
            break
        if not isinstance(no, (int, float)) or not name:
            continue
        year_month_raw = row[3] if len(row) > 3 else None
        year_month = _yy_mm_to_iso(str(year_month_raw)) if year_month_raw else None
        company = _normalize_company(row[6]) if len(row) > 6 and row[6] else "미상"
        amount = clean_amount(row[8]) if len(row) > 8 else None
        if amount is None:
            continue
        records.append(
            {
                "company": company,
                "person": str(name).strip(),
                "year_month": year_month or sheet_year_month,
                "category": "retirement",
                "amount": amount,
            }
        )
    return records


def extract_hr_costs(path, filename):
    """엑셀 파일에서 건강/연금/퇴직공제 인당 납부 레코드를 추출한다.

    파일이 어떤 양식인지는 시트 이름 패턴으로 자동 판별한다.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        records = []
        matched_any = False

        for sheet_name in wb.sheetnames:
            m = PREMIUM_SHEET_RE.match(sheet_name)
            if m:
                matched_any = True
                category = PREMIUM_CATEGORY_MAP[m.group(2)]
                year_month = _yy_mm_to_iso(m.group(3))
                records.extend(_parse_premium_sheet(wb[sheet_name], category, year_month))
                continue

            m = RETIREMENT_SHEET_RE.match(sheet_name)
            if m:
                matched_any = True
                sheet_year_month = _yy_mm_to_iso(f"{m.group(1)}.{m.group(2)}")
                records.extend(_parse_retirement_sheet(wb[sheet_name], sheet_year_month))

        if not matched_any:
            raise ValueError(
                f"지원하지 않는 형식의 엑셀 파일입니다: {filename} "
                "(보험료 납부_단위공사별 또는 퇴직공제부금 납부 신고 내역 양식만 지원합니다.)"
            )
        return records
    finally:
        wb.close()


def _empty_totals():
    return {cat: 0 for cat in CATEGORY_ORDER}


def aggregate_by_company(records):
    totals = {}
    for r in records:
        totals.setdefault(r["company"], _empty_totals())[r["category"]] += r["amount"]
    result = []
    for company, cat_totals in sorted(totals.items(), key=lambda kv: kv[0] or ""):
        result.append({"company": company, **cat_totals, "total": sum(cat_totals.values())})
    return result


def aggregate_by_person(records):
    totals = {}
    for r in records:
        key = (r["company"], r["person"])
        totals.setdefault(key, _empty_totals())[r["category"]] += r["amount"]
    result = []
    for (company, person), cat_totals in sorted(totals.items(), key=lambda kv: (kv[0][0] or "", kv[0][1] or "")):
        result.append({"company": company, "person": person, **cat_totals, "total": sum(cat_totals.values())})
    return result


def aggregate_by_month(records):
    totals = {}
    for r in records:
        ym = r["year_month"] or "미상"
        totals.setdefault(ym, _empty_totals())[r["category"]] += r["amount"]
    result = []
    for ym, cat_totals in sorted(totals.items()):
        result.append({"year_month": ym, **cat_totals, "total": sum(cat_totals.values())})
    return result


def aggregate_by_company_month(records):
    """DB 저장용: 성명을 제외하고 (업체, 연월, 구분)별로 합산한 평평한 행 목록을 반환한다.
    금액이 0인 구분은 행 자체를 만들지 않는다."""
    totals = {}
    for r in records:
        key = (r["company"], r["year_month"] or "미상")
        totals.setdefault(key, _empty_totals())[r["category"]] += r["amount"]
    result = []
    for (company, ym), cat_totals in sorted(totals.items(), key=lambda kv: (kv[0][0] or "", kv[0][1] or "")):
        for cat in CATEGORY_ORDER:
            if cat_totals[cat]:
                result.append({"company": company, "year_month": ym, "category": cat, "amount": cat_totals[cat]})
    return result


def aggregate_detail(records):
    """(업체, 성명, 연월) 단위로 구분별 금액을 합쳐 상세표를 만든다."""
    totals = {}
    for r in records:
        key = (r["company"], r["person"], r["year_month"] or "미상")
        totals.setdefault(key, _empty_totals())[r["category"]] += r["amount"]
    result = []
    for (company, person, ym), cat_totals in sorted(totals.items(), key=lambda kv: (kv[0][0] or "", kv[0][2] or "", kv[0][1] or "")):
        result.append(
            {"company": company, "person": person, "year_month": ym, **cat_totals, "total": sum(cat_totals.values())}
        )
    return result


def build_export_workbook(records):
    wb = Workbook()
    cat_headers = [CATEGORY_LABELS[c] for c in CATEGORY_ORDER]

    ws = wb.active
    ws.title = "상세"
    ws.append(["업체명", "성명", "연월", *cat_headers, "합계"])
    for row in aggregate_detail(records):
        ws.append([row["company"], row["person"], row["year_month"], *[row[c] for c in CATEGORY_ORDER], row["total"]])

    ws2 = wb.create_sheet("업체별 집계")
    ws2.append(["업체명", *cat_headers, "합계"])
    for row in aggregate_by_company(records):
        ws2.append([row["company"], *[row[c] for c in CATEGORY_ORDER], row["total"]])

    ws3 = wb.create_sheet("개인별 집계")
    ws3.append(["업체명", "성명", *cat_headers, "합계"])
    for row in aggregate_by_person(records):
        ws3.append([row["company"], row["person"], *[row[c] for c in CATEGORY_ORDER], row["total"]])

    ws4 = wb.create_sheet("월별 집계")
    ws4.append(["연월", *cat_headers, "합계"])
    for row in aggregate_by_month(records):
        ws4.append([row["year_month"], *[row[c] for c in CATEGORY_ORDER], row["total"]])

    return wb
